from __future__ import annotations

from copy import copy
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app_info import APP_EXCEL_TITLE, APP_NAME, APP_SCAN_UNION_HTML_NAME, APP_WORKBOOK_NAME

OUT = ROOT / "dist" / APP_WORKBOOK_NAME
SCAN_UNION_HTML = APP_SCAN_UNION_HTML_NAME

INPUT = APP_NAME
CALC = "Calc"
UV_POWER = "UVPower"
UV_DB = "UVData"
UV_VIEW = "UVPlot"
STRUCTURE_DATA = "StructureData"
PATTERN_3D = "Pattern3D"
CUT_DATA = "CutData"
SCAN_COVERAGE_DATA = "ScanCoverage2D"
CUT_VIEW = "CutPlot"
ENV_3D = "Envelope3D"
ELEMENT_NORM = "ElementNorm"
PATTERN_WIRE = "PatternWire"
ENV_WIRE = "EnvelopeWire"
NOTE = "Note"

BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_YELLOW = "FFF2CC"
GRID = "D9E2F3"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = INPUT
    ws_calc = wb.create_sheet(CALC)
    ws_uv_power = wb.create_sheet(UV_POWER)
    ws_uv_db = wb.create_sheet(UV_DB)
    ws_structure = wb.create_sheet(STRUCTURE_DATA)
    ws_cut_data = wb.create_sheet(CUT_DATA)
    ws_scan_coverage = wb.create_sheet(SCAN_COVERAGE_DATA)
    ws_elem_norm = wb.create_sheet(ELEMENT_NORM)

    build_dashboard(ws_dashboard, uv_n=61)
    build_calculations(ws_calc)
    build_element_norm(ws_elem_norm, theta_n=121, phi_n=181)
    build_uv_power(ws_uv_power, n=61)
    build_uv_db(ws_uv_db, n=61)
    build_structure_data(ws_structure, max_elements=1024)
    build_cut_data(ws_cut_data, n=121)
    build_scan_coverage_data(ws_scan_coverage, n=121, scan_n=31)
    build_dashboard_plots(ws_dashboard, uv_n=61, cut_n=121)

    for ws in (ws_calc, ws_uv_power, ws_uv_db, ws_structure, ws_cut_data, ws_scan_coverage, ws_elem_norm):
        ws.sheet_state = "hidden"

    for ws in wb.worksheets:
        style_sheet(ws)
    format_dashboard(ws_dashboard, uv_n=61)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.active = 0
    wb.save(OUT)
    print(OUT)


def render_core_preview_images() -> dict[str, Path]:
    import sys

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))

    from core.array_factor import calc_uv_pattern
    from core.element_pattern import make_element_pattern
    from core.envelope import compute_current_3d_envelope
    from core.geometry import BeamParams, derive_params, get_mode_settings

    image_dir = OUT.parent / "excel_core_preview"
    image_dir.mkdir(parents=True, exist_ok=True)
    pattern_path = image_dir / "pattern3d_core.png"
    envelope_path = image_dir / "envelope3d_core.png"

    params = BeamParams(
        frequency_ghz=5.6,
        nx=2,
        ny=8,
        dx_m=0.7,
        dy_m=0.1,
        ax_m=0.7,
        ay_m=0.1,
        efficiency=0.7,
        element_power_w=1_000_000.0,
        s0_w_cm2=20.0,
        scan_x_deg=0.0,
        scan_y_deg=0.0,
        calc_mode="standard",
        use_element_pattern=True,
    )
    params, derived = derive_params(params)
    elem = make_element_pattern(params, derived.wavelength_m)
    settings = get_mode_settings("standard")

    uv = calc_uv_pattern(params, derived, elem, settings.n_uv, settings.pattern_floor_db)
    u = np.asarray(uv["U"], dtype=float)
    v = np.asarray(uv["V"], dtype=float)
    z = np.asarray(uv["pattern_db"], dtype=float)
    visible = np.asarray(uv["visible"], dtype=bool) & np.isfinite(z)
    theta_x = np.degrees(np.arcsin(np.clip(u, -1.0, 1.0)))
    theta_y = np.degrees(np.arcsin(np.clip(v, -1.0, 1.0)))
    z_masked = np.ma.masked_where(~visible, z)
    fig = plt.figure(figsize=(7.8, 5.0), dpi=140)
    ax = fig.add_subplot(111, projection="3d")
    surf = ax.plot_surface(
        theta_x,
        theta_y,
        z_masked,
        cmap="viridis",
        vmin=settings.pattern_floor_db,
        vmax=0.0,
        linewidth=0.12,
        antialiased=True,
        rstride=2,
        cstride=2,
    )
    ax.set_title("3D Pattern (dB)")
    ax.set_xlabel("theta_x (deg)")
    ax.set_ylabel("theta_y (deg)")
    ax.set_zlabel("dB")
    ax.set_zlim(settings.pattern_floor_db, 0.0)
    ax.view_init(elev=28, azim=-58)
    fig.colorbar(surf, ax=ax, shrink=0.62, pad=0.08, label="dB")
    fig.tight_layout()
    fig.savefig(pattern_path, bbox_inches="tight")
    plt.close(fig)

    env = compute_current_3d_envelope(params, derived, elem, settings)
    u_env = np.asarray(env["U"], dtype=float)
    v_env = np.asarray(env["V"], dtype=float)
    r = np.asarray(env["r_env_m"], dtype=float)
    mask = np.isfinite(r) & (r > 0.0)
    u_masked = np.ma.masked_where(~mask, u_env)
    v_masked = np.ma.masked_where(~mask, v_env)
    r_masked = np.ma.masked_where(~mask, r)
    fig = plt.figure(figsize=(7.8, 5.0), dpi=140)
    ax = fig.add_subplot(111, projection="3d")
    surf = ax.plot_surface(
        u_masked,
        v_masked,
        r_masked,
        cmap="viridis",
        vmin=0.0,
        vmax=float(np.nanmax(r)) if np.any(np.isfinite(r)) else 1.0,
        linewidth=0.10,
        antialiased=True,
        rstride=2,
        cstride=2,
    )
    ax.set_title("S=S0 Envelope r(u,v)")
    ax.set_xlabel("u")
    ax.set_ylabel("v")
    ax.set_zlabel("r (m)")
    ax.set_xlim(-1.0, 1.0)
    ax.set_ylim(-1.0, 1.0)
    if np.any(np.isfinite(r)):
        ax.set_zlim(0.0, float(np.nanmax(r)) * 1.05)
    ax.view_init(elev=28, azim=-58)
    fig.colorbar(surf, ax=ax, shrink=0.62, pad=0.08, label="r (m)")
    fig.tight_layout()
    fig.savefig(envelope_path, bbox_inches="tight")
    plt.close(fig)
    return {"pattern": pattern_path, "envelope": envelope_path}


def build_dashboard(ws, uv_n: int) -> None:
    ws["A1"] = APP_EXCEL_TITLE
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:AO1")
    ws["A2"] = "只修改黄色单元格；本页保留二维图表。自由旋转三维扫描并集请打开同目录 HTML。"
    ws.merge_cells("A2:AO2")

    rows = [
        ("频率 f", 5.6, "GHz", "输入"),
        ("Nx", 2, "", "输入"),
        ("Ny", 8, "", "输入"),
        ("dx", 0.7, "m", "输入"),
        ("dy", 0.1, "m", "输入"),
        ("ax", 0.7, "m", "输入"),
        ("ay", 0.1, "m", "输入"),
        ("效率 eta", 0.7, "", "输入"),
        ("每单元功率", 1_000_000.0, "W", "输入"),
        ("S0", 20.0, "W/cm2", "输入"),
        ("x-z切面扫描角", 0.0, "deg", "输入"),
        ("y-z切面扫描角", 0.0, "deg", "输入"),
        ("最大扫描角模式", "自动", "", "自动/手动"),
        ("手动X最大扫描角", 8.0, "deg", "手动模式有效"),
        ("手动Y最大扫描角", 8.0, "deg", "手动模式有效"),
        ("使用矩形口径单元方向图", 1, "1/0", "1=使用, 0=不使用"),
        ("方向图显示下限", -40.0, "dB", "输入"),
    ]
    headers = ["参数", "值", "单位", "备注"]
    ws["A3"] = "参数输入"
    ws["A3"].font = Font(bold=True, color=BLUE, size=13)
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    for row_idx, row in enumerate(rows, 5):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 2).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        ws.cell(row_idx, 2).protection = Protection(locked=False)

    mode_validation = DataValidation(type="list", formula1='"自动,手动"', allow_blank=False)
    bool_validation = DataValidation(type="list", formula1='"1,0"', allow_blank=False)
    ws.add_data_validation(mode_validation)
    ws.add_data_validation(bool_validation)
    mode_validation.add(ws["B17"])
    bool_validation.add(ws["B20"])

    ws["F3"] = "自动计算结果"
    ws["F3"].font = Font(bold=True, color=BLUE, size=13)
    result_headers = ["项目", "值", "单位", "说明"]
    for col, header in enumerate(result_headers, 6):
        ws.cell(4, col, header)
    result_rows = [
        ("lambda", "='Calc'!B4", "m", "波长"),
        ("Dx", "='Calc'!B10", "m", "阵列口径 x"),
        ("Dy", "='Calc'!B11", "m", "阵列口径 y"),
        ("Rff", "='Calc'!B13", "m", "远场边界"),
        ("HPBWx", "='Calc'!B14", "deg", "经验半功率波束宽度"),
        ("HPBWy", "='Calc'!B15", "deg", "经验半功率波束宽度"),
        ("单元归一化增益", "='Calc'!B25", "", "与 EXE element_gain_norm 对齐"),
        ("最大扫描角X", "='Calc'!B16", "deg", "自动/手动"),
        ("最大扫描角Y", "='Calc'!B17", "deg", "自动/手动"),
        ("theta", "='Calc'!B8", "deg", "当前扫描俯仰角"),
        ("phi", "='Calc'!B9", "deg", "当前扫描方位角"),
        ("当前扫描损失", "='Calc'!B22", "dB", "矩形口径方向图 * cos(theta)"),
        ("相位差X", "='Calc'!B18", "deg", "相邻单元"),
        ("相位差Y", "='Calc'!B19", "deg", "相邻单元"),
        ("总输入功率", "='Calc'!B20", "W", "Nx*Ny*单元功率"),
    ]
    for row_idx, row in enumerate(result_rows, 5):
        for col_idx, value in enumerate(row, 6):
            ws.cell(row_idx, col_idx, value)

    ws["F20"] = "三维扫描并集"
    ws["G20"] = "打开交互HTML"
    ws["G20"].hyperlink = SCAN_UNION_HTML
    ws["G20"].style = "Hyperlink"
    ws["H20"] = "同目录文件"
    ws["I20"] = SCAN_UNION_HTML

    ws["K3"] = "u-v方向图 (dB)"
    ws["K3"].font = Font(bold=True, color=BLUE, size=13)
    ws.merge_cells("K3:AO3")
    ws["K4"] = "横轴 u，纵轴 v；颜色越亮表示归一化方向图越高"
    ws.merge_cells("K4:AO4")
    uv_view_n = 31
    stride = max(1, (uv_n - 1) // (uv_view_n - 1))
    start_row = 6
    start_col = 11
    ws.cell(start_row - 1, start_col - 1, "v/u")
    for i in range(uv_view_n):
        src_idx = i * stride
        col = start_col + i
        row = start_row + i
        source_col = get_column_letter(2 + src_idx)
        source_row = 3 + src_idx
        if i % 5 == 0 or i in (0, uv_view_n - 1):
            ws.cell(start_row - 1, col, f"='{UV_DB}'!{source_col}$2")
            ws.cell(row, start_col - 1, f"='{UV_DB}'!$A{source_row}")
        for j in range(uv_view_n):
            source_data_col = get_column_letter(2 + j * stride)
            ws.cell(row, start_col + j, f"='{UV_DB}'!{source_data_col}{source_row}")
    heatmap_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + uv_view_n - 1)}{start_row + uv_view_n - 1}"
    ws.conditional_formatting.add(
        heatmap_range,
        ColorScaleRule(
            start_type="num",
            start_value=-40,
            start_color="440154",
            mid_type="num",
            mid_value=-15,
            mid_color="22A884",
            end_type="num",
            end_value=0,
            end_color="FDE725",
        ),
    )
    ws.freeze_panes = "A5"


def build_inputs(ws) -> None:
    ws["A1"] = APP_EXCEL_TITLE
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:D1")
    ws["A2"] = "用户只需要修改黄色单元格；Excel 会自动计算并刷新图表。"
    ws.merge_cells("A2:D2")

    rows = [
        ("频率 f", 10.0, "GHz", "输入"),
        ("Nx", 8, "", "输入"),
        ("Ny", 8, "", "输入"),
        ("dx", 0.1, "m", "输入"),
        ("dy", 0.1, "m", "输入"),
        ("ax", 0.1, "m", "输入"),
        ("ay", 0.1, "m", "输入"),
        ("效率 eta", 0.7, "", "输入"),
        ("每单元功率", 1_000_000.0, "W", "输入"),
        ("S0", 20.0, "W/cm2", "输入"),
        ("x-z切面扫描角", 8.0, "deg", "输入"),
        ("y-z切面扫描角", 8.0, "deg", "输入"),
        ("最大扫描角模式", "自动", "", "自动/手动"),
        ("手动X最大扫描角", 8.0, "deg", "手动模式有效"),
        ("手动Y最大扫描角", 8.0, "deg", "手动模式有效"),
        ("使用矩形口径单元方向图", 1, "1/0", "1=使用, 0=不使用"),
        ("方向图显示下限", -40.0, "dB", "输入"),
    ]
    headers = ["参数", "值", "单位", "备注"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    for row_idx, row in enumerate(rows, 5):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 2).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        ws.cell(row_idx, 2).protection = Protection(locked=False)

    mode_validation = DataValidation(type="list", formula1='"自动,手动"', allow_blank=False)
    bool_validation = DataValidation(type="list", formula1='"1,0"', allow_blank=False)
    ws.add_data_validation(mode_validation)
    ws.add_data_validation(bool_validation)
    mode_validation.add(ws["B17"])
    bool_validation.add(ws["B20"])
    ws.freeze_panes = "A5"


def build_calculations(ws) -> None:
    ws["A1"] = "自动计算结果"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:D1")
    headers = ["项目", "值", "单位", "公式说明"]
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)

    rows = [
        ("lambda", f"=0.3/'{INPUT}'!$B$5", "m", "lambda = c/f"),
        ("u0", f"=SIN(RADIANS('{INPUT}'!$B$15))", "", "u0 = sin(scanX)"),
        ("v0", f"=SIN(RADIANS('{INPUT}'!$B$16))", "", "v0 = sin(scanY)"),
        ("w0", "=SQRT(MAX(0,1-B5^2-B6^2))", "", "w0 = sqrt(1-u0^2-v0^2)"),
        ("theta", "=DEGREES(ASIN(SQRT(B5^2+B6^2)))", "deg", "theta = asin(sqrt(u0^2+v0^2))"),
        ("phi", "=IF(B5^2+B6^2<1E-24,0,DEGREES(ATAN2(B5,B6)))", "deg", "phi = atan2(v0,u0)"),
        ("Dx", f"=('{INPUT}'!$B$6-1)*'{INPUT}'!$B$8+'{INPUT}'!$B$10", "m", "阵列口径"),
        ("Dy", f"=('{INPUT}'!$B$7-1)*'{INPUT}'!$B$9+'{INPUT}'!$B$11", "m", "阵列口径"),
        ("Aperture", "=MAX(B10,B11)", "m", "最大口径"),
        ("Rff", "=2*B12^2/B4", "m", "远场边界"),
        ("HPBWx", "=2*DEGREES(ASIN(MIN(1,0.442946*B4/B10)))", "deg", "经验半功率波束宽度"),
        ("HPBWy", "=2*DEGREES(ASIN(MIN(1,0.442946*B4/B11)))", "deg", "经验半功率波束宽度"),
        ("经验最大扫描角X", f"=IF('{INPUT}'!$B$17=\"手动\",ABS('{INPUT}'!$B$18),DEGREES(ASIN(MIN(1,0.442946*B4/'{INPUT}'!$B$10))))", "deg", "自动/手动"),
        ("经验最大扫描角Y", f"=IF('{INPUT}'!$B$17=\"手动\",ABS('{INPUT}'!$B$19),DEGREES(ASIN(MIN(1,0.442946*B4/'{INPUT}'!$B$11))))", "deg", "自动/手动"),
        ("相邻单元相位差X", f"=-360*'{INPUT}'!$B$8/B4*B5", "deg", "相邻阵元"),
        ("相邻单元相位差Y", f"=-360*'{INPUT}'!$B$9/B4*B6", "deg", "相邻阵元"),
        ("总输入功率", f"='{INPUT}'!$B$6*'{INPUT}'!$B$7*'{INPUT}'!$B$13", "W", "Nx*Ny*单元功率"),
        ("S0", f"='{INPUT}'!$B$14*10000", "W/m2", "W/cm2 -> W/m2"),
        ("当前扫描损失", scan_loss_formula(), "dB", "矩形口径方向图 * cos(theta)"),
        ("u-v方向图峰值功率", f"=MAX('{UV_POWER}'!B3:BJ63)", "", "用于归一化"),
        ("三维方向图峰值功率", "=B23", "", "与u-v方向图共用归一化峰值"),
        ("单元归一化增益", f"=IF('{INPUT}'!$B$20=1,'{ELEMENT_NORM}'!$B$1,2)", "", "EXE element_gain_norm"),
    ]
    for row_idx, row in enumerate(rows, 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = "A4"


def build_uv_power(ws, n: int) -> None:
    ws["A1"] = "u-v方向图原始功率"
    step = 2.0 / (n - 1)
    for i in range(n):
        ws.cell(2, i + 2, -1.0 + i * step)
        ws.cell(i + 3, 1, -1.0 + i * step)
    for row in range(3, n + 3):
        for col in range(2, n + 2):
            ws.cell(row, col, f"={raw_pattern_formula(f'{get_column_letter(col)}$2', f'$A{row}')}")


def build_element_norm(ws, theta_n: int, phi_n: int) -> None:
    ws["A1"] = "element_gain_norm"
    ws["B1"] = f"=IF('{INPUT}'!$B$20=1,4*PI()/MAX($B$4,1E-300),2)"
    ws["A3"] = "dtheta"
    ws["B3"] = f"={(1.5707963267948966 / (theta_n - 1)):.16g}"
    ws["A4"] = "integral"
    last_col = get_column_letter(phi_n + 1)
    last_row = theta_n + 7
    ws["B4"] = f"=$B$3*$B$5*SUM(B8:{last_col}{last_row})"
    ws["A5"] = "dphi"
    ws["B5"] = f"={(6.283185307179586 / (phi_n - 1)):.16g}"
    ws["A7"] = "theta/phi"

    for j in range(phi_n):
        col = j + 2
        phi = -3.141592653589793 + j * 6.283185307179586 / (phi_n - 1)
        ws.cell(7, col, phi)
    for i in range(theta_n):
        row = i + 8
        theta = i * 1.5707963267948966 / (theta_n - 1)
        theta_weight = 0.5 if i in (0, theta_n - 1) else 1.0
        ws.cell(row, 1, theta)
        for j in range(phi_n):
            col = j + 2
            phi_weight = 0.5 if j in (0, phi_n - 1) else 1.0
            weight = theta_weight * phi_weight
            theta_ref = f"$A{row}"
            phi_ref = f"{get_column_letter(col)}$7"
            ux_term = f"'{INPUT}'!$B$10/'{CALC}'!$B$4*SIN({theta_ref})*COS({phi_ref})"
            uy_term = f"'{INPUT}'!$B$11/'{CALC}'!$B$4*SIN({theta_ref})*SIN({phi_ref})"
            sx = f"IF(ABS({ux_term})<1E-12,1,(SIN(PI()*({ux_term}))/MAX(ABS(PI()*({ux_term})),1E-12))^2)"
            sy = f"IF(ABS({uy_term})<1E-12,1,(SIN(PI()*({uy_term}))/MAX(ABS(PI()*({uy_term})),1E-12))^2)"
            ws.cell(row, col, f"={weight:.1f}*({sx})*({sy})*MAX(COS({theta_ref}),0)*SIN({theta_ref})")


def build_uv_db(ws, n: int) -> None:
    ws["A1"] = "u-v方向图 dB 数据"
    for i in range(n):
        ws.cell(2, i + 2, f"='{UV_POWER}'!{get_column_letter(i + 2)}$2")
        ws.cell(i + 3, 1, f"='{UV_POWER}'!$A{i + 3}")
    for row in range(3, n + 3):
        for col in range(2, n + 2):
            src = f"'{UV_POWER}'!{get_column_letter(col)}{row}"
            u_ref = f"'{UV_POWER}'!{get_column_letter(col)}$2"
            v_ref = f"'{UV_POWER}'!$A{row}"
            ws.cell(row, col, f"=IF(({u_ref})^2+({v_ref})^2>1,\"\",MAX('{INPUT}'!$B$21,10*LOG10(MAX({src},1E-300)/MAX('{CALC}'!$B$23,1E-300))))")


def build_structure_data(ws, max_elements: int) -> None:
    ws["A1"] = "even_x"
    ws["B1"] = "even_y"
    ws["C1"] = "odd_x"
    ws["D1"] = "odd_y"
    frame_headers = ["left_x", "left_y", "right_x", "right_y", "bottom_x", "bottom_y", "top_x", "top_y"]
    for col, header in enumerate(frame_headers, 6):
        ws.cell(1, col, header)
    for idx in range(max_elements):
        row = idx + 2
        ix = f"MOD({idx},'{INPUT}'!$B$6)"
        iy = f"INT({idx}/'{INPUT}'!$B$6)"
        active = f"{idx}<'{INPUT}'!$B$6*'{INPUT}'!$B$7"
        x_expr = f"(({ix})-('{INPUT}'!$B$6-1)/2)*'{INPUT}'!$B$8"
        y_expr = f"(({iy})-('{INPUT}'!$B$7-1)/2)*'{INPUT}'!$B$9"
        even = f"MOD(({ix})+({iy}),2)=0"
        ws.cell(row, 1, f"=IF(AND({active},{even}),{x_expr},NA())")
        ws.cell(row, 2, f"=IF(AND({active},{even}),{y_expr},NA())")
        ws.cell(row, 3, f"=IF(AND({active},NOT({even})),{x_expr},NA())")
        ws.cell(row, 4, f"=IF(AND({active},NOT({even})),{y_expr},NA())")

    frame_series = [
        (6, [("=-'Calc'!$B$10/2", "=-'Calc'!$B$11/2"), ("=-'Calc'!$B$10/2", "='Calc'!$B$11/2")]),
        (8, [("='Calc'!$B$10/2", "=-'Calc'!$B$11/2"), ("='Calc'!$B$10/2", "='Calc'!$B$11/2")]),
        (10, [("=-'Calc'!$B$10/2", "=-'Calc'!$B$11/2"), ("='Calc'!$B$10/2", "=-'Calc'!$B$11/2")]),
        (12, [("=-'Calc'!$B$10/2", "='Calc'!$B$11/2"), ("='Calc'!$B$10/2", "='Calc'!$B$11/2")]),
    ]
    for col, points in frame_series:
        for row_offset, (x_formula, y_formula) in enumerate(points, 2):
            ws.cell(row_offset, col, x_formula)
            ws.cell(row_offset, col + 1, y_formula)


def build_uv_view(ws, n: int) -> None:
    ws["A1"] = "u-v方向图热力图（dB）"
    ws["A2"] = "修改 Input 页后，Excel 自动刷新。红色星号位置请参考 Calc 页中的 u0/v0。"
    for i in range(n):
        ws.cell(4, i + 2, f"='{UV_DB}'!{get_column_letter(i + 2)}$2")
        ws.cell(i + 5, 1, f"='{UV_DB}'!$A{i + 3}")
    for row in range(5, n + 5):
        for col in range(2, n + 2):
            ws.cell(row, col, f"='{UV_DB}'!{get_column_letter(col)}{row - 2}")
    data_range = f"B5:{get_column_letter(n + 1)}{n + 4}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="num",
            start_value=-40,
            start_color="440154",
            mid_type="num",
            mid_value=-15,
            mid_color="22A884",
            end_type="num",
            end_value=0,
            end_color="FDE725",
        ),
    )
    ws.freeze_panes = "B5"


def build_pattern_3d(ws, n: int) -> None:
    ws["A1"] = "三维方向图 Pattern (dB)"
    ws["A2"] = "下方曲面图由本表公式数据生成，横纵轴为角度，竖轴为归一化方向图 dB。"
    min_angle, max_angle = -80.0, 80.0
    step = (max_angle - min_angle) / (n - 1)
    for i in range(n):
        angle = min_angle + i * step
        ws.cell(4, i + 2, angle)
        ws.cell(i + 5, 1, angle)
    for row in range(5, n + 5):
        for col in range(2, n + 2):
            u = f"SIN(RADIANS({get_column_letter(col)}$4))"
            v = f"SIN(RADIANS($A{row}))"
            raw = raw_pattern_formula(u, v)
            ws.cell(row, col, f"=MAX('{INPUT}'!$B$21,10*LOG10(MAX({raw},1E-300)/MAX('{CALC}'!$B$24,1E-300)))")

    ws.freeze_panes = "B5"


def build_cut_data(ws, n: int) -> None:
    ws["A1"] = "二维切面 S=S0 包络数据"
    headers = ["alpha_deg", "xz_r_m", "xz_x_m", "xz_z_m", "yz_r_m", "yz_y_m", "yz_z_m"]
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    min_angle, max_angle = -89.0, 89.0
    step = (max_angle - min_angle) / (n - 1)
    for row_idx in range(n):
        row = row_idx + 4
        ws.cell(row, 1, min_angle + row_idx * step)
        alpha = f"$A{row}"
        u_x = f"SIN(RADIANS({alpha}))"
        v_x = "0"
        u_y = "0"
        v_y = f"SIN(RADIANS({alpha}))"
        ws.cell(row, 2, envelope_radius_formula(u_x, v_x))
        ws.cell(row, 3, f"=B{row}*{u_x}")
        ws.cell(row, 4, f"=B{row}*SQRT(MAX(0,1-({u_x})^2))")
        ws.cell(row, 5, envelope_radius_formula(u_y, v_y))
        ws.cell(row, 6, f"=E{row}*{v_y}")
        ws.cell(row, 7, f"=E{row}*SQRT(MAX(0,1-({v_y})^2))")
    ws.freeze_panes = "A4"


def build_scan_coverage_data(ws, n: int, scan_n: int) -> None:
    ws["A1"] = "二维切面所有扫描角度下等EIRP覆盖数据"
    headers = [
        "alpha_deg",
        "xz_union_r_m",
        "xz_union_x_m",
        "xz_union_z_m",
        "yz_union_r_m",
        "yz_union_y_m",
        "yz_union_z_m",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    x_scan_start_col = 8
    y_scan_start_col = x_scan_start_col + scan_n
    for scan_idx in range(scan_n):
        ws.cell(2, x_scan_start_col + scan_idx, f"x_scan_{scan_idx}")
        ws.cell(2, y_scan_start_col + scan_idx, f"y_scan_{scan_idx}")
    min_angle, max_angle = -89.0, 89.0
    step = (max_angle - min_angle) / (n - 1)
    for row_idx in range(n):
        row = row_idx + 4
        ws.cell(row, 1, min_angle + row_idx * step)
        alpha = f"$A{row}"
        u_x = f"SIN(RADIANS({alpha}))"
        v_y = f"SIN(RADIANS({alpha}))"
        for scan_idx in range(scan_n):
            t = -1.0 + 2.0 * scan_idx / (scan_n - 1)
            u0_scan = f"SIN(RADIANS(({t:.12g})*'Calc'!$B$16))"
            v0_scan = f"SIN(RADIANS(({t:.12g})*'Calc'!$B$17))"
            ws.cell(row, x_scan_start_col + scan_idx, f"={raw_pattern_formula_at_scan(u_x, '0', u0_scan, '0')}")
            ws.cell(row, y_scan_start_col + scan_idx, f"={raw_pattern_formula_at_scan('0', v_y, '0', v0_scan)}")
        x_max = f"{get_column_letter(x_scan_start_col)}{row}:{get_column_letter(x_scan_start_col + scan_n - 1)}{row}"
        y_max = f"{get_column_letter(y_scan_start_col)}{row}:{get_column_letter(y_scan_start_col + scan_n - 1)}{row}"
        ws.cell(row, 2, radius_from_raw_formula(f"MAX({x_max})"))
        ws.cell(row, 3, f"=B{row}*{u_x}")
        ws.cell(row, 4, f"=B{row}*SQRT(MAX(0,1-({u_x})^2))")
        ws.cell(row, 5, radius_from_raw_formula(f"MAX({y_max})"))
        ws.cell(row, 6, f"=E{row}*{v_y}")
        ws.cell(row, 7, f"=E{row}*SQRT(MAX(0,1-({v_y})^2))")
    ws.freeze_panes = "A4"


def build_cut_view(ws, n: int) -> None:
    ws["A1"] = "二维固定切面 S=S0 包络"
    chart_x = ScatterChart()
    chart_x.title = "x-z固定切面"
    chart_x.x_axis.title = "x (m)"
    chart_x.y_axis.title = "z (m)"
    chart_x.height = 12
    chart_x.width = 16
    xvalues = Reference(ws.parent[CUT_DATA], min_col=3, min_row=4, max_row=n + 3)
    yvalues = Reference(ws.parent[CUT_DATA], min_col=4, min_row=4, max_row=n + 3)
    chart_x.series.append(Series(yvalues, xvalues, title="x-z"))
    ws.add_chart(chart_x, "A3")

    chart_y = ScatterChart()
    chart_y.title = "y-z固定切面"
    chart_y.x_axis.title = "y (m)"
    chart_y.y_axis.title = "z (m)"
    chart_y.height = 12
    chart_y.width = 16
    xvalues = Reference(ws.parent[CUT_DATA], min_col=6, min_row=4, max_row=n + 3)
    yvalues = Reference(ws.parent[CUT_DATA], min_col=7, min_row=4, max_row=n + 3)
    chart_y.series.append(Series(yvalues, xvalues, title="y-z"))
    ws.add_chart(chart_y, "J3")


def build_env_3d(ws, v_n: int, u_n: int) -> None:
    ws["A1"] = "三维包络距离面 r(m)"
    ws["A2"] = "u-v网格与 EXE 标准模式三维包络采样一致，显示 S=S0 包络距离面。"
    uv_max = 0.999
    for i in range(u_n):
        ws.cell(4, i + 2, -uv_max + 2.0 * uv_max * i / (u_n - 1))
    for i in range(v_n):
        ws.cell(i + 5, 1, -uv_max + 2.0 * uv_max * i / (v_n - 1))
    for row in range(5, v_n + 5):
        for col in range(2, u_n + 2):
            u = f"{get_column_letter(col)}$4"
            v = f"$A{row}"
            ws.cell(row, col, envelope_radius_formula(u, v))
    ws.freeze_panes = "B5"


def build_note(ws) -> None:
    ws["A1"] = "说明"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    notes = [
        "这是独立 Excel 版本，用户在 Input 页修改黄色单元格即可自动计算。",
        "方向图、二维切面和三维包络距离面均由 Excel 公式生成，图表会随公式数据刷新。",
        "本 Excel 采用远场阵因子/矩形口径单元方向图公式，适合快速参数交流和客户演示。",
        "精确扫描并集三维覆盖涉及大量扫描状态取并集，纯公式 Excel 会非常慢；该功能请使用 EXE/APK。",
        "如 Excel 打开后未自动刷新，请按 F9 或在“公式 -> 计算选项”中选择“自动”。",
    ]
    for row, text in enumerate(notes, 3):
        ws.cell(row, 1, text)
    ws.column_dimensions["A"].width = 120


def build_pattern_wire(ws, n: int, step: int) -> None:
    ws["A1"] = "三维方向图投影线框数据"
    indices = _selected_indices(n, step)
    series_count = 0

    def add_series(title: str, point_formulas: list[tuple[str, str]]) -> None:
        nonlocal series_count
        col = series_count * 2 + 1
        ws.cell(1, col, title)
        ws.cell(2, col, "x_proj")
        ws.cell(2, col + 1, "y_proj")
        for i, (x_formula, y_formula) in enumerate(point_formulas, 3):
            ws.cell(i, col, x_formula)
            ws.cell(i, col + 1, y_formula)
        series_count += 1

    for row_idx in indices:
        src_row = 5 + row_idx
        y_angle = f"'{PATTERN_3D}'!$A{src_row}"
        points: list[tuple[str, str]] = []
        for col_idx in range(n):
            src_col = get_column_letter(2 + col_idx)
            x_angle = f"'{PATTERN_3D}'!{src_col}$4"
            db = f"'{PATTERN_3D}'!{src_col}{src_row}"
            x_proj = f"={x_angle}-0.55*({y_angle})"
            y_proj = f"=({db}-'{INPUT}'!$B$21)*1.8+0.18*({x_angle}+{y_angle})"
            points.append((x_proj, y_proj))
        add_series(f"row_{row_idx}", points)

    for col_idx in indices:
        src_col = get_column_letter(2 + col_idx)
        x_angle = f"'{PATTERN_3D}'!{src_col}$4"
        points = []
        for row_idx in range(n):
            src_row = 5 + row_idx
            y_angle = f"'{PATTERN_3D}'!$A{src_row}"
            db = f"'{PATTERN_3D}'!{src_col}{src_row}"
            x_proj = f"={x_angle}-0.55*({y_angle})"
            y_proj = f"=({db}-'{INPUT}'!$B$21)*1.8+0.18*({x_angle}+{y_angle})"
            points.append((x_proj, y_proj))
        add_series(f"col_{col_idx}", points)

    ws["B1"] = series_count


def build_env_wire(ws, v_n: int, u_n: int, row_step: int, col_step: int) -> None:
    ws["A1"] = "三维包络投影线框数据"
    series_count = 0
    max_points = max(v_n, u_n)

    def projected_formulas(src_row: int, src_col: int) -> tuple[str, str]:
        u_ref = f"'{ENV_3D}'!{get_column_letter(src_col)}$4"
        v_ref = f"'{ENV_3D}'!$A{src_row}"
        r_ref = f"'{ENV_3D}'!{get_column_letter(src_col)}{src_row}"
        visible = f"({u_ref})^2+({v_ref})^2<=0.998001"
        x_proj = f"=IF({visible},100*(({u_ref})-0.55*({v_ref})),\"\")"
        y_proj = f"=IF({visible},({r_ref})+18*(({u_ref})+({v_ref})),\"\")"
        return x_proj, y_proj

    def add_series(title: str, point_formulas: list[tuple[str, str]]) -> None:
        nonlocal series_count
        col = series_count * 2 + 1
        ws.cell(1, col, title)
        ws.cell(2, col, "x_proj")
        ws.cell(2, col + 1, "y_proj")
        for i in range(max_points):
            row = i + 3
            if i < len(point_formulas):
                ws.cell(row, col, point_formulas[i][0])
                ws.cell(row, col + 1, point_formulas[i][1])
            else:
                ws.cell(row, col, "")
                ws.cell(row, col + 1, "")
        series_count += 1

    for v_idx in _selected_indices(v_n, row_step):
        src_row = 5 + v_idx
        points = [projected_formulas(src_row, 2 + u_idx) for u_idx in range(u_n)]
        add_series(f"v_{v_idx}", points)

    for u_idx in _selected_indices(u_n, col_step):
        src_col = 2 + u_idx
        points = [projected_formulas(5 + v_idx, src_col) for v_idx in range(v_n)]
        add_series(f"u_{u_idx}", points)

    ws["B1"] = series_count


def build_dashboard_plots(ws, uv_n: int, cut_n: int) -> None:
    ws["A37"] = "结构示意 / 二维等 EIRP / 二维扫描覆盖"
    ws["A37"].font = Font(bold=True, color=BLUE, size=13)
    ws["A38"] = "六个图块：结构示意、u-v方向图、x-z/y-z当前等EIRP、x-z/y-z所有扫描角覆盖。隐藏工作表仅作为公式和图表数据源。"
    ws.merge_cells("A38:AO38")
    ws["A38"].font = Font(color="666666", italic=True)

    structure = ScatterChart()
    structure.title = "结构示意图"
    structure.x_axis.title = "y (m)"
    structure.y_axis.title = "x (m)"
    structure.legend = None
    structure.height = 11.5
    structure.width = 13.2
    even_x = Reference(ws.parent[STRUCTURE_DATA], min_col=1, min_row=2, max_row=1025)
    even_y = Reference(ws.parent[STRUCTURE_DATA], min_col=2, min_row=2, max_row=1025)
    odd_x = Reference(ws.parent[STRUCTURE_DATA], min_col=3, min_row=2, max_row=1025)
    odd_y = Reference(ws.parent[STRUCTURE_DATA], min_col=4, min_row=2, max_row=1025)
    s_even = Series(even_y, even_x, title="阵元A")
    s_odd = Series(odd_y, odd_x, title="阵元B")
    for series, color in ((s_even, "2F75B5"), (s_odd, "F4B183")):
        try:
            series.marker.symbol = "square"
            series.marker.size = 7
            series.graphicalProperties.line.noFill = True
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = "FFFFFF"
        except Exception:
            pass
    structure.series.append(s_even)
    structure.series.append(s_odd)
    for col in (6, 8, 10, 12):
        frame_x = Reference(ws.parent[STRUCTURE_DATA], min_col=col, min_row=2, max_row=3)
        frame_y = Reference(ws.parent[STRUCTURE_DATA], min_col=col + 1, min_row=2, max_row=3)
        s_frame = Series(frame_y, frame_x, title="阵列口径")
        try:
            s_frame.marker.symbol = "none"
            s_frame.graphicalProperties.line.solidFill = "C00000"
            s_frame.graphicalProperties.line.width = 18000
            s_frame.smooth = False
        except Exception:
            pass
        structure.series.append(s_frame)
    ws.add_chart(structure, "A39")

    chart_x = ScatterChart()
    chart_x.title = "x-z 等EIRP切面"
    chart_x.x_axis.title = "z (m)"
    chart_x.y_axis.title = "x (m)"
    chart_x.height = 11.5
    chart_x.width = 13.2
    xvalues = Reference(ws.parent[CUT_DATA], min_col=3, min_row=4, max_row=cut_n + 3)
    yvalues = Reference(ws.parent[CUT_DATA], min_col=4, min_row=4, max_row=cut_n + 3)
    s_x = Series(yvalues, xvalues, title="x-z")
    try:
        s_x.marker.symbol = "none"
        s_x.graphicalProperties.line.solidFill = "4472C4"
        s_x.graphicalProperties.line.width = 18000
    except Exception:
        pass
    chart_x.series.append(s_x)
    ws.add_chart(chart_x, "K39")

    chart_y = ScatterChart()
    chart_y.title = "y-z 等EIRP切面"
    chart_y.x_axis.title = "z (m)"
    chart_y.y_axis.title = "y (m)"
    chart_y.height = 11.5
    chart_y.width = 13.2
    xvalues = Reference(ws.parent[CUT_DATA], min_col=6, min_row=4, max_row=cut_n + 3)
    yvalues = Reference(ws.parent[CUT_DATA], min_col=7, min_row=4, max_row=cut_n + 3)
    s_y = Series(yvalues, xvalues, title="y-z")
    try:
        s_y.marker.symbol = "none"
        s_y.graphicalProperties.line.solidFill = "4472C4"
        s_y.graphicalProperties.line.width = 18000
    except Exception:
        pass
    chart_y.series.append(s_y)
    ws.add_chart(chart_y, "U39")

    ws["A62"] = "所有允许扫描角下二维覆盖包络"
    ws["A62"].font = Font(bold=True, color=BLUE, size=13)
    ws["A63"] = "x-z图按 ±最大扫描角X 对 scanX 取并集；y-z图按 ±最大扫描角Y 对 scanY 取并集。"
    ws.merge_cells("A63:AO63")
    ws["A63"].font = Font(color="666666", italic=True)

    chart_x_union = ScatterChart()
    chart_x_union.title = "x-z 全扫描等EIRP覆盖"
    chart_x_union.x_axis.title = "z (m)"
    chart_x_union.y_axis.title = "x (m)"
    chart_x_union.height = 11.5
    chart_x_union.width = 15.0
    xvalues = Reference(ws.parent[SCAN_COVERAGE_DATA], min_col=3, min_row=4, max_row=cut_n + 3)
    yvalues = Reference(ws.parent[SCAN_COVERAGE_DATA], min_col=4, min_row=4, max_row=cut_n + 3)
    s_xu = Series(yvalues, xvalues, title="x-z all scan")
    try:
        s_xu.marker.symbol = "none"
        s_xu.graphicalProperties.line.solidFill = "C00000"
        s_xu.graphicalProperties.line.width = 18000
    except Exception:
        pass
    chart_x_union.series.append(s_xu)
    ws.add_chart(chart_x_union, "A64")

    chart_y_union = ScatterChart()
    chart_y_union.title = "y-z 全扫描等EIRP覆盖"
    chart_y_union.x_axis.title = "z (m)"
    chart_y_union.y_axis.title = "y (m)"
    chart_y_union.height = 11.5
    chart_y_union.width = 15.0
    xvalues = Reference(ws.parent[SCAN_COVERAGE_DATA], min_col=6, min_row=4, max_row=cut_n + 3)
    yvalues = Reference(ws.parent[SCAN_COVERAGE_DATA], min_col=7, min_row=4, max_row=cut_n + 3)
    s_yu = Series(yvalues, xvalues, title="y-z all scan")
    try:
        s_yu.marker.symbol = "none"
        s_yu.graphicalProperties.line.solidFill = "C00000"
        s_yu.graphicalProperties.line.width = 18000
    except Exception:
        pass
    chart_y_union.series.append(s_yu)
    ws.add_chart(chart_y_union, "Q64")


def raw_pattern_formula(u_expr: str, v_expr: str) -> str:
    return raw_pattern_formula_at_scan(u_expr, v_expr, f"'{CALC}'!$B$5", f"'{CALC}'!$B$6")


def raw_pattern_formula_at_scan(u_expr: str, v_expr: str, u0_expr: str, v0_expr: str) -> str:
    u = f"({u_expr})"
    v = f"({v_expr})"
    u0 = f"({u0_expr})"
    v0 = f"({v0_expr})"
    kx = f"(PI()*'{INPUT}'!$B$8/'{CALC}'!$B$4*({u}-{u0}))"
    ky = f"(PI()*'{INPUT}'!$B$9/'{CALC}'!$B$4*({v}-{v0}))"
    sx = f"('{INPUT}'!$B$10/'{CALC}'!$B$4*{u})"
    sy = f"('{INPUT}'!$B$11/'{CALC}'!$B$4*{v})"
    afx = f"MIN(1,IF(ABS(SIN({kx}))<1E-9,1,(SIN('{INPUT}'!$B$6*{kx})/('{INPUT}'!$B$6*MAX(ABS(SIN({kx})),1E-9)))^2))"
    afy = f"MIN(1,IF(ABS(SIN({ky}))<1E-9,1,(SIN('{INPUT}'!$B$7*{ky})/('{INPUT}'!$B$7*MAX(ABS(SIN({ky})),1E-9)))^2))"
    ex = f"MIN(1,IF(ABS({sx})<1E-12,1,(SIN(PI()*{sx})/MAX(ABS(PI()*{sx}),1E-12))^2))"
    ey = f"MIN(1,IF(ABS({sy})<1E-12,1,(SIN(PI()*{sy})/MAX(ABS(PI()*{sy}),1E-12))^2))"
    elem = f"IF('{INPUT}'!$B$20=1,({ex})*({ey}),1)"
    return f"IF({u}^2+{v}^2>1,0,({afx})*({afy})*({elem}))"


def envelope_radius_formula(u_expr: str, v_expr: str) -> str:
    raw = raw_pattern_formula(u_expr, v_expr)
    return radius_from_raw_formula(raw)


def radius_from_raw_formula(raw_expr: str) -> str:
    return (
        f"=SQRT(MAX(0,'{INPUT}'!$B$12*'{INPUT}'!$B$13*"
        f"'{CALC}'!$B$25*('{INPUT}'!$B$6*'{INPUT}'!$B$7)^2*({raw_expr})/(4*PI()*MAX('{CALC}'!$B$21,1E-300))))"
    )


def scan_loss_formula() -> str:
    sx = f"('{INPUT}'!$B$10/B4*B5)"
    sy = f"('{INPUT}'!$B$11/B4*B6)"
    ex = f"MIN(1,IF(ABS({sx})<1E-12,1,(SIN(PI()*{sx})/MAX(ABS(PI()*{sx}),1E-12))^2))"
    ey = f"MIN(1,IF(ABS({sy})<1E-12,1,(SIN(PI()*{sy})/MAX(ABS(PI()*{sy}),1E-12))^2))"
    return f"=10*LOG10(MAX(1E-300,({ex})*({ey})*B7))"


def _selected_indices(n: int, step: int) -> list[int]:
    indices = list(range(0, int(n), max(1, int(step))))
    last = int(n) - 1
    if not indices or indices[-1] != last:
        indices.append(last)
    return indices


def format_dashboard(ws, uv_n: int) -> None:
    widths = {
        "A": 20,
        "B": 13,
        "C": 9,
        "D": 18,
        "E": 2,
        "F": 18,
        "G": 14,
        "H": 9,
        "I": 18,
        "J": 6,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    uv_view_n = 31
    start_col = 11
    for col_idx in range(start_col, start_col + uv_view_n):
        ws.column_dimensions[get_column_letter(col_idx)].width = 2.6
    for row_idx in range(6, 6 + uv_view_n):
        ws.row_dimensions[row_idx].height = 13

    for cell in ws["A1:AO1"][0]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color="FFFFFF", bold=True, size=16)
    for row in (4,):
        for col in range(1, 10):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.font = Font(bold=True)
    for col in range(11, 42):
        ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
    for row in range(6, 37):
        ws.cell(row, 10).alignment = Alignment(horizontal="center", vertical="center")
    for row in range(5, 22):
        ws.cell(row, 2).number_format = "0.0000"
    for row in range(5, 19):
        ws.cell(row, 7).number_format = "0.0000"
    ws["G18"].number_format = "0.000E+00"
    ws.sheet_view.showGridLines = False


def style_sheet(ws) -> None:
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row in (1, 3, 4) and cell.value is not None:
                font = copy(cell.font)
                font.bold = True
                cell.font = font
    for col in range(1, min(ws.max_column, 12) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    main()
