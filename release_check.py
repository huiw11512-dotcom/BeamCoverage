from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
import re
import subprocess
import sys
from zipfile import ZipFile

from app_info import (
    APP_APK_NAME,
    APP_EXCEL_TITLE,
    APP_EXE_NAME,
    APP_NAME,
    APP_RELEASE_DIR_NAME,
    APP_RELEASE_ZIP_NAME,
    APP_SCAN_UNION_HTML_NAME,
    APP_SCAN_UNION_HTML_TITLE,
    APP_SPEC_NAME,
    APP_SYNC_ZIP_NAME,
    APP_TITLE,
    APP_VERSION,
    APP_WORKBOOK_NAME,
)


ROOT = Path(__file__).resolve().parent
EXCEL_COMPANION_HTML = APP_SCAN_UNION_HTML_NAME
EXPECTED_FILES = [APP_EXE_NAME, APP_APK_NAME, APP_WORKBOOK_NAME, EXCEL_COMPANION_HTML, "README.txt", "CHANGELOG.md"]
OLD_BRAND_MARKERS = (
    "相控阵功率密度包络",
    "扫描覆盖预测软件",
    "BeamCoverage V",
    "BeamCoverage_V",
    "BeamCoverage Excel V",
)


def _windows_version_text(app_version: str) -> str:
    parts = str(app_version).strip().split(".")
    if len(parts) == 3:
        parts.append("0")
    if len(parts) != 4 or any(not part.isdigit() for part in parts):
        raise ValueError(f"APP_VERSION must have 3 or 4 numeric parts for Windows version info, got {app_version!r}")
    return ".".join(parts)


EXPECTED_WINDOWS_VERSION = _windows_version_text(APP_VERSION)
EXPECTED_VERSION_INFO = {
    "FileDescription": APP_NAME,
    "ProductName": APP_NAME,
    "FileVersion": EXPECTED_WINDOWS_VERSION,
    "ProductVersion": EXPECTED_WINDOWS_VERSION,
    "OriginalFilename": APP_EXE_NAME,
    "InternalName": APP_NAME,
}


def main() -> int:
    args = _parse_args()
    release_dir = args.release_dir.resolve()
    zip_path = args.zip_path.resolve()
    sync_zip_path = args.sync_zip_path.resolve()
    smoke_output = args.smoke_output.resolve()

    _check_required_files(release_dir)
    _check_release_readme_documentation(release_dir)
    _check_active_build_inputs()
    check_version_resource_source()
    _check_version_info(release_dir / APP_EXE_NAME)
    _check_exe_smoke(release_dir / APP_EXE_NAME, smoke_output)
    _check_excel_workbook(release_dir)
    _check_apk_signature(release_dir / APP_APK_NAME)
    _check_zip(zip_path, release_dir, label="release zip")
    _check_zip(sync_zip_path, release_dir, label="sync zip")

    print("PASS release package check.")
    print(f"PASS release dir: {release_dir}")
    print(f"PASS release zip: {zip_path}")
    print(f"PASS sync zip: {sync_zip_path}")
    print(r"For full source + release validation, run: python tools\run_validation.py --release")
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=f"Validate {APP_NAME} release artifacts.")
    parser.add_argument(
        "--release-dir",
        type=Path,
        default=ROOT / "dist" / APP_RELEASE_DIR_NAME,
        help=f"Directory containing {APP_EXE_NAME}/{APP_APK_NAME}/{APP_WORKBOOK_NAME}/HTML/README.txt.",
    )
    parser.add_argument(
        "--zip-path",
        type=Path,
        default=ROOT / "dist" / APP_RELEASE_ZIP_NAME,
        help="Release zip to validate.",
    )
    parser.add_argument(
        "--sync-zip-path",
        type=Path,
        default=ROOT / "dist" / APP_SYNC_ZIP_NAME,
        help="Convenience/sync zip to validate. Defaults to the user-facing BeamCoverage.zip.",
    )
    parser.add_argument(
        "--smoke-output",
        type=Path,
        default=ROOT / "dist" / "_validation" / "smoke_release_check.json",
        help="Temporary JSON output written by packaged EXE smoke test.",
    )
    return parser.parse_args()


def _check_required_files(release_dir: Path) -> None:
    if not release_dir.exists():
        raise AssertionError(f"Release directory does not exist: {release_dir}")
    missing = [name for name in EXPECTED_FILES if not (release_dir / name).exists()]
    if missing:
        raise AssertionError(f"Release directory is missing files: {missing}")
    empty = [name for name in EXPECTED_FILES if (release_dir / name).stat().st_size <= 0]
    if empty:
        raise AssertionError(f"Release directory has empty files: {empty}")
    print("PASS required release files exist.")


def _check_release_readme_documentation(release_dir: Path) -> None:
    from tools.check_docs import check_release_docs
    from tools.build_release import _default_release_readme

    readme_path = release_dir / "README.txt"
    check_release_docs(readme_path)
    actual = readme_path.read_text(encoding="utf-8").replace("\r\n", "\n")
    expected = _default_release_readme().replace("\r\n", "\n")
    if actual != expected:
        raise AssertionError(
            "Release README.txt is not the current generated template from tools/build_release.py. "
            "Run python tools\\build_release.py to regenerate the release package."
        )
    _reject_old_brand_markers("release README.txt", actual)
    print("PASS release README documentation.")


def _check_active_build_inputs() -> None:
    versioned_specs = sorted(path.name for path in ROOT.glob(f"{APP_NAME}_V*.spec"))
    if versioned_specs:
        raise AssertionError(f"Old versioned PyInstaller specs are still active: {versioned_specs}")
    current_spec = ROOT / APP_SPEC_NAME
    if not current_spec.exists():
        raise AssertionError(f"Current PyInstaller spec is missing: {current_spec}")

    build_dir = ROOT / "build"
    versioned_builds = sorted(path.name for path in build_dir.glob(f"{APP_NAME}_V*") if path.is_dir()) if build_dir.exists() else []
    if versioned_builds:
        raise AssertionError(f"Old versioned PyInstaller build directories are still active: {versioned_builds}")
    print(f"PASS active build inputs use {APP_NAME} only.")


def check_version_resource_source(*, verbose: bool = True) -> None:
    source = ROOT / "version_info.txt"
    if not source.exists():
        raise AssertionError(f"Windows version resource file is missing: {source}")
    text = source.read_text(encoding="utf-8")
    required_fragments = [
        f"filevers=({', '.join(EXPECTED_WINDOWS_VERSION.split('.'))})",
        f"prodvers=({', '.join(EXPECTED_WINDOWS_VERSION.split('.'))})",
        f"StringStruct('FileDescription', '{APP_NAME}')",
        f"StringStruct('FileVersion', '{EXPECTED_WINDOWS_VERSION}')",
        f"StringStruct('InternalName', '{APP_NAME}')",
        f"StringStruct('OriginalFilename', '{APP_EXE_NAME}')",
        f"StringStruct('ProductName', '{APP_NAME}')",
        f"StringStruct('ProductVersion', '{EXPECTED_WINDOWS_VERSION}')",
    ]
    compact = re.sub(r"\s+", "", text)
    _reject_old_brand_markers("version_info.txt", text)
    missing = []
    for fragment in required_fragments:
        if re.sub(r"\s+", "", fragment) not in compact:
            missing.append(fragment)
    if missing:
        raise AssertionError(f"version_info.txt is not synchronized with app_info.py: missing {missing}")
    if verbose:
        print("PASS version_info.txt matches app_info.py.")


def _check_version_info(exe_path: Path) -> None:
    actual = _read_windows_version_info(exe_path)
    mismatches = {
        key: {"expected": expected, "actual": actual.get(key)}
        for key, expected in EXPECTED_VERSION_INFO.items()
        if actual.get(key) != expected
    }
    if mismatches:
        raise AssertionError(f"Unexpected EXE version info: {mismatches}")
    print(f"PASS EXE version info is {APP_NAME}.")


def _read_windows_version_info(exe_path: Path) -> dict[str, str]:
    command = "\n".join(
        [
            "$ErrorActionPreference = 'Stop'",
            f"$v = (Get-Item -LiteralPath {_ps_quote(str(exe_path))}).VersionInfo",
            "[pscustomobject]@{",
            "  FileDescription = $v.FileDescription",
            "  ProductName = $v.ProductName",
            "  FileVersion = $v.FileVersion",
            "  ProductVersion = $v.ProductVersion",
            "  OriginalFilename = $v.OriginalFilename",
            "  InternalName = $v.InternalName",
            "} | ConvertTo-Json -Compress",
        ]
    )
    proc = subprocess.run(
        ["powershell", "-NoProfile", "-Command", command],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if proc.returncode != 0:
        raise AssertionError(f"PowerShell version-info query failed: {proc.stderr.strip()}")
    return json.loads(proc.stdout)


def _ps_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _check_exe_smoke(exe_path: Path, smoke_output: Path) -> None:
    smoke_output.parent.mkdir(parents=True, exist_ok=True)
    if smoke_output.exists():
        smoke_output.unlink()
    env = os.environ.copy()
    env["QT_QPA_PLATFORM"] = "offscreen"
    env["OPENBLAS_NUM_THREADS"] = "1"
    env["OMP_NUM_THREADS"] = "1"
    env["MKL_NUM_THREADS"] = "1"
    proc = subprocess.run(
        [str(exe_path), "--smoke-test", "--smoke-output", str(smoke_output)],
        check=False,
        env=env,
        timeout=120,
    )
    if proc.returncode != 0:
        raise AssertionError(f"Packaged EXE smoke test failed with exit code {proc.returncode}")
    payload = json.loads(smoke_output.read_text(encoding="utf-8"))
    title = str(payload.get("title", ""))
    if title != APP_TITLE:
        raise AssertionError(f"Packaged EXE title mismatch: expected {APP_TITLE!r}, got {title!r}")
    _reject_old_brand_markers("packaged EXE title", title)
    required = {
        "ok": True,
        "title": APP_TITLE,
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "menu_localized": True,
        "plot_toolbars_ok": True,
        "surface_gap_preserve_ok": True,
        "default_report_figures_ok": True,
        "current_method_metadata_ok": True,
        "union_report_figures_ok": True,
        "union_cuts_ok": True,
        "union_cut_high_resolution_ok": True,
        "union_method_metadata_ok": True,
        "union_timing_breakdown_ok": True,
        "scan_angle_cache_reuse_ok": True,
        "scan_angle_full_cache_reuse_ok": True,
        "settings_cache_key_distinguishes_resolution_ok": True,
        "project_paths_resolved_ok": True,
        "project_save_relative_paths_ok": True,
        "radian_pattern_loaded_ok": True,
        "metadata_vector_pattern_loaded_ok": True,
        "multifrequency_pattern_loaded_ok": True,
        "imported_layout_requires_custom_ok": True,
        "inactive_layout_save_relative_ok": True,
        "inactive_layout_file_ignored_ok": True,
        "inactive_pattern_file_ignored_ok": True,
        "string_false_pattern_ignored_ok": True,
        "missing_import_status_labels_ok": True,
        "missing_project_import_status_ok": True,
        "excel_report_inactive_pattern_label_ok": True,
        "excel_report_inactive_layout_label_ok": True,
        "generated_geometry_overlap_rejected": True,
        "generated_geometry_calculate_disabled": True,
        "generated_geometry_single_axis_ok": True,
        "app_icon_loaded": True,
        "window_icon_loaded": True,
        "icon_sizes_ok": True,
        "export_templates_ok": True,
        "exported_element_pattern_abs_phase_template_vector": True,
        "exported_element_pattern_abs_phase_template_has_phase": True,
        "default_result_metadata_ok": True,
        "imported_element_pattern_loaded": True,
        "imported_element_pattern_calculated": True,
        "imported_element_pattern_has_phase": True,
        "imported_element_pattern_status_has_metadata": True,
        "imported_element_pattern_covers_visible_edge": False,
        "imported_vector_element_pattern_loaded": True,
        "imported_vector_element_pattern_calculated": True,
        "imported_vector_element_pattern_status_has_metadata": True,
        "imported_array_layout_loaded": True,
        "imported_array_layout_calculated": True,
        "imported_array_layout_status_has_metadata": True,
        "imported_result_metadata_ok": True,
        "imported_vector_result_metadata_ok": True,
        "imported_array_overlap_rejected": True,
        "imported_array_overlap_calculate_disabled": True,
    }
    mismatches = {
        key: {"expected": expected, "actual": payload.get(key)}
        for key, expected in required.items()
        if payload.get(key) != expected
    }
    if mismatches:
        raise AssertionError(f"Packaged EXE smoke payload mismatch: {mismatches}")
    theta_max = payload.get("imported_element_pattern_theta_max_deg")
    if not isinstance(theta_max, (int, float)) or not 79.5 <= float(theta_max) <= 80.5:
        raise AssertionError(f"Packaged smoke theta_max metadata is unexpected: {theta_max!r}")
    status_text = str(payload.get("imported_element_pattern_status_text", ""))
    if "\u03b8max=80.0" not in status_text or "\u672a\u8986\u76d690" not in status_text:
        raise AssertionError(f"Packaged smoke status text lacks theta coverage metadata: {status_text!r}")
    vector_status_text = str(payload.get("imported_vector_element_pattern_status_text", ""))
    if "\u77e2\u91cfE\u03b8/E\u03c6" not in vector_status_text:
        raise AssertionError(f"Packaged smoke vector status text lacks vector metadata: {vector_status_text!r}")
    if int(payload.get("exported_array_template_point_count", 0)) <= 0:
        raise AssertionError(f"Packaged smoke exported array template count is invalid: {payload.get('exported_array_template_point_count')!r}")
    if payload.get("exported_element_pattern_template_vector") is not True:
        raise AssertionError("Packaged smoke exported element pattern template was not recognized as vector data.")
    if payload.get("exported_element_pattern_abs_phase_template_vector") is not True:
        raise AssertionError("Packaged smoke exported Abs/Phase element pattern template was not recognized as vector data.")
    if payload.get("exported_element_pattern_abs_phase_template_has_phase") is not True:
        raise AssertionError("Packaged smoke exported Abs/Phase element pattern template was not recognized as phase data.")
    app_icon_sizes = payload.get("app_icon_sizes")
    window_icon_sizes = payload.get("window_icon_sizes")
    if not _has_large_icon_size(app_icon_sizes) or not _has_large_icon_size(window_icon_sizes):
        raise AssertionError(
            "Packaged smoke did not expose usable application/window icon sizes: "
            f"app={app_icon_sizes!r}, window={window_icon_sizes!r}"
        )
    overlap_warning = str(payload.get("imported_array_overlap_warning_text", ""))
    if "\u53e3\u5f84\u91cd\u53e0" not in overlap_warning:
        raise AssertionError(f"Packaged smoke overlap warning did not mention footprint overlap: {overlap_warning!r}")
    layout_status = str(payload.get("imported_array_layout_status_text", ""))
    if "\u6709\u6548\u5355\u5143" not in layout_status or "\u603b\u529f\u7387" not in layout_status or "\u53e3\u5f84" not in layout_status:
        raise AssertionError(f"Packaged smoke layout status lacks imported-array metadata: {layout_status!r}")
    default_result = (
        payload.get("default_result_array_source_text"),
        payload.get("default_result_active_elements_text"),
        payload.get("default_result_element_model_text"),
    )
    if default_result != ("\u77e9\u5f62\u6392\u5e03", "16", "\u77e9\u5f62\u89e3\u6790\u53e3\u5f84"):
        raise AssertionError(f"Packaged smoke default result metadata is unexpected: {default_result!r}")
    imported_result = (
        payload.get("imported_result_array_source_text"),
        payload.get("imported_result_active_elements_text"),
        payload.get("imported_result_element_model_text"),
    )
    if imported_result != ("\u5bfc\u5165\u5750\u6807CSV", "5", "\u5bfc\u5165\u8fdc\u573a\u65b9\u5411\u56fe"):
        raise AssertionError(f"Packaged smoke imported result metadata is unexpected: {imported_result!r}")
    if payload.get("imported_vector_result_element_model_text") != "\u5bfc\u5165\u8fdc\u573a\u77e2\u91cfE\u03b8/E\u03c6":
        raise AssertionError(
            "Packaged smoke vector element model metadata is unexpected: "
            f"{payload.get('imported_vector_result_element_model_text')!r}"
        )
    print("PASS packaged EXE smoke test.")


def _has_large_icon_size(value: object) -> bool:
    if not isinstance(value, list):
        return False
    for item in value:
        if (
            isinstance(item, list)
            and len(item) == 2
            and isinstance(item[0], int)
            and isinstance(item[1], int)
            and item[0] >= 64
            and item[1] >= 64
        ):
            return True
    return False


def _check_apk_signature(apk_path: Path) -> None:
    apksigner = _find_apksigner()
    if not apksigner:
        print("SKIP APK signature check: apksigner not found.")
        return
    env = os.environ.copy()
    bundled_jdk = Path(r"E:\AndroidDev\jdk\jdk-17")
    if "JAVA_HOME" not in env and (bundled_jdk / "bin" / "java.exe").exists():
        env["JAVA_HOME"] = str(bundled_jdk)
    proc = subprocess.run(
        [str(apksigner), "-JXmx256M", "verify", "--verbose", str(apk_path)],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
    )
    if proc.returncode != 0:
        raise AssertionError(f"APK signature check failed: {proc.stdout}\n{proc.stderr}")
    if "Verified using" not in proc.stdout:
        raise AssertionError(f"APK signature output did not contain verification lines: {proc.stdout}")
    print("PASS APK signature verification.")


def _find_apksigner() -> Path | None:
    candidates = [
        Path(r"E:\AndroidDev\android-sdk\build-tools\34.0.0\apksigner.bat"),
    ]
    for env_name in ("ANDROID_HOME", "ANDROID_SDK_ROOT"):
        root = os.environ.get(env_name)
        if root:
            candidates.extend(sorted(Path(root).glob("build-tools/*/apksigner.bat"), reverse=True))
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _check_excel_workbook(release_dir: Path) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - report environment issues clearly.
        raise AssertionError(f"openpyxl is required for release Excel validation: {exc}") from exc

    workbook_path = release_dir / APP_WORKBOOK_NAME
    html_path = release_dir / EXCEL_COMPANION_HTML
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    if APP_NAME not in wb.sheetnames:
        raise AssertionError(f"Release workbook is missing {APP_NAME} sheet: {wb.sheetnames}")
    ws = wb[APP_NAME]
    title = str(ws["A1"].value or "")
    if title != APP_EXCEL_TITLE:
        raise AssertionError(f"Release workbook has unexpected title: {title!r}")
    _reject_old_brand_markers("release workbook title", title)

    link = ws["G20"].hyperlink
    target = link.target if link else None
    if target != EXCEL_COMPANION_HTML:
        raise AssertionError(f"Release workbook 3D companion link is wrong: {target!r}")
    if not html_path.exists() or html_path.stat().st_size <= 0:
        raise AssertionError(f"Release workbook companion HTML is missing or empty: {html_path}")
    html = html_path.read_text(encoding="utf-8", errors="ignore")
    if APP_SCAN_UNION_HTML_TITLE not in html or "Plotly.newPlot" not in html:
        raise AssertionError("Release workbook companion HTML does not look like the interactive scan-union plot.")
    _reject_old_brand_markers("release workbook companion HTML", html)
    print("PASS Excel workbook and companion HTML.")


def _reject_old_brand_markers(label: str, text: str) -> None:
    hits = [marker for marker in OLD_BRAND_MARKERS if marker in text]
    if hits:
        raise AssertionError(f"{label} still contains old prototype/verbose brand markers: {hits}")


def _check_zip(zip_path: Path, release_dir: Path, *, label: str) -> None:
    if not zip_path.exists():
        raise AssertionError(f"{label} does not exist: {zip_path}")
    with ZipFile(zip_path, "r") as zf:
        infos = zf.infolist()
        names = [info.filename for info in infos]
        if sorted(names) != sorted(EXPECTED_FILES):
            raise AssertionError(f"Unexpected {label} entries: {names}")
        zip_sizes = {info.filename: info.file_size for info in infos}
    mismatches = {
        name: {"release": (release_dir / name).stat().st_size, "zip": zip_sizes.get(name)}
        for name in EXPECTED_FILES
        if (release_dir / name).stat().st_size != zip_sizes.get(name)
    }
    if mismatches:
        raise AssertionError(f"{label} entry sizes do not match release files: {mismatches}")
    print(f"PASS {label} entries match release directory.")


if __name__ == "__main__":
    raise SystemExit(main())
