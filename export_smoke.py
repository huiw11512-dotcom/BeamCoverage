from __future__ import annotations

import csv
from pathlib import Path
import shutil
import sys

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.array_factor import calc_uv_pattern
from core.element_pattern import load_imported_element_pattern, make_element_pattern
from core.envelope import compute_2d_cuts, compute_current_3d_envelope
from core.geometry import BeamParams, derive_params, get_mode_settings, load_imported_array_layout
from core.near_field import export_near_field_projected_far_field_pattern, load_imported_element_near_field
from core.scan_union import compute_scan_union_envelope_3d
from export.csv_export import export_2d_cuts, export_current_3d, export_scan_union, export_structure, export_uv_pattern
from export.csv_templates import (
    export_current_array_layout_template,
    export_element_near_field_vector_template,
    export_element_pattern_abs_phase_template,
    export_element_pattern_vector_template,
)
from export.excel_export import export_excel_report
from export.png_export import export_figure_png


def main() -> int:
    out_dir = ROOT / "export_smoke_output"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir()

    params = BeamParams(nx=4, ny=4, element_power_w=1.0e4, calc_mode="fast")
    params, derived = derive_params(params)
    elem = make_element_pattern(params, derived.wavelength_m)
    settings0 = get_mode_settings("fast")
    settings = settings0.__class__(
        **{
            **settings0.__dict__,
            "n_alpha_2d": 21,
            "n_range_2d": 16,
            "theta_3d_n": 8,
            "phi_3d_n": 12,
            "n_range_3d": 12,
            "n_uv": 31,
            "scan_union_step_deg": 12.0,
            "scan_union_theta_n": 8,
            "scan_union_phi_n": 12,
            "chunk_size": 256,
            "scan_block_size": 8,
        }
    )
    cuts = compute_2d_cuts(params, derived, elem, settings)
    env = compute_current_3d_envelope(params, derived, elem, settings)
    uv = calc_uv_pattern(params, derived, elem, settings.n_uv, settings.pattern_floor_db)
    union = compute_scan_union_envelope_3d(params, derived, elem, settings)
    union_timing = union.get("timings", {})
    for timing_key in ("scan_union_3d_s", "scan_union_2d_cuts_s", "scan_union_compute_s"):
        if not isinstance(union_timing, dict) or timing_key not in union_timing:
            raise RuntimeError(f"scan union did not include timing breakdown {timing_key}.")

    export_structure(
        {
            "x_m": derived.xe_m.ravel(),
            "y_m": derived.ye_m.ravel(),
            "z_m": derived.xe_m.ravel() * 0.0,
            "power_w": derived.xe_m.ravel() * 0.0 + params.element_power_w,
        },
        out_dir / "structure.csv",
    )
    export_2d_cuts(cuts, out_dir / "2d_cuts.csv", params.s0_w_cm2)
    cuts_csv = out_dir / "2d_cuts.csv"
    with cuts_csv.open("r", newline="", encoding="utf-8-sig") as handle:
        cut_rows = list(csv.DictReader(handle))
    if not cut_rows or "envelope_method" not in cut_rows[0] or cut_rows[0]["envelope_method"] != "near_field_sampled_far_field_extrapolated":
        raise RuntimeError("2d_cuts.csv did not export current-envelope method metadata.")

    current_3d_csv = out_dir / "current_3d.csv"
    export_current_3d(env, current_3d_csv, params.s0_w_cm2)
    with current_3d_csv.open("r", newline="", encoding="utf-8-sig") as handle:
        current_rows = list(csv.DictReader(handle))
    if not current_rows or "envelope_method" not in current_rows[0] or current_rows[0]["envelope_method"] != "near_field_sampled_far_field_extrapolated":
        raise RuntimeError("current_3d.csv did not export current-envelope method metadata.")
    export_uv_pattern(uv, out_dir / "uv.csv")
    scan_union_csv = out_dir / "scan_union.csv"
    export_scan_union(union, scan_union_csv)
    with scan_union_csv.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    datasets = {row.get("dataset") for row in rows}
    cut_names = {row.get("cut_name") for row in rows if row.get("dataset") == "2d_union_cut"}
    if "3d_union" not in datasets or "2d_union_cut" not in datasets:
        raise RuntimeError("scan_union.csv did not include both 3D union rows and 2D union cut rows.")
    if {"scan_union_xz_fixed_cut", "scan_union_yz_fixed_cut"} - cut_names:
        raise RuntimeError(f"scan_union.csv is missing fixed scan-union cut rows: {cut_names}")
    if not rows or "envelope_method" not in rows[0] or rows[0]["envelope_method"] != "far_field_coefficient_union":
        raise RuntimeError("scan_union.csv did not export scan-union method metadata.")
    union_cut_rows = [row for row in rows if row.get("dataset") == "2d_union_cut"]
    if not union_cut_rows or any(row.get("envelope_method") != "near_field_sampled_far_field_scan_union" for row in union_cut_rows):
        raise RuntimeError("scan_union.csv did not export near-field scan-union cut method metadata.")
    array_template = out_dir / "array_layout_template.csv"
    pattern_template = out_dir / "element_pattern_vector_template.csv"
    abs_phase_pattern_template = out_dir / "element_pattern_abs_phase_template.csv"
    near_field_template = out_dir / "element_near_field_template.csv"
    near_field_projected_pattern = out_dir / "element_near_field_projected_far_field.csv"
    export_current_array_layout_template(derived, array_template)
    exported_layout = load_imported_array_layout(array_template, params.element_power_w)
    if exported_layout.point_count != int(derived.element_x_m.size):
        raise RuntimeError("Exported current array layout template did not round-trip active element count.")
    export_element_pattern_vector_template(pattern_template)
    exported_pattern = load_imported_element_pattern(pattern_template)
    if not exported_pattern.has_vector_components:
        raise RuntimeError("Exported element pattern template was not recognized as ETheta/EPhi vector data.")
    export_element_pattern_abs_phase_template(abs_phase_pattern_template)
    exported_abs_phase_pattern = load_imported_element_pattern(abs_phase_pattern_template)
    if not exported_abs_phase_pattern.has_vector_components or not exported_abs_phase_pattern.has_phase:
        raise RuntimeError("Exported Abs/Phase element pattern template was not recognized as vector phase data.")
    export_element_near_field_vector_template(near_field_template)
    exported_near_field = load_imported_element_near_field(near_field_template, params.frequency_ghz)
    if not exported_near_field.has_vector_components or not exported_near_field.has_phase:
        raise RuntimeError("Exported near-field template was not recognized as complex Ex/Ey/Ez vector data.")
    projection_info = export_near_field_projected_far_field_pattern(
        near_field_template,
        near_field_projected_pattern,
        frequency_ghz=params.frequency_ghz,
    )
    if int(projection_info["output_rows"]) <= 0:
        raise RuntimeError("Near-field projected far-field export did not report output rows.")
    projected_pattern = load_imported_element_pattern(near_field_projected_pattern, params.frequency_ghz)
    if not projected_pattern.has_vector_components or projected_pattern.point_count <= 0:
        raise RuntimeError("Near-field projected far-field CSV did not round-trip as vector far-field data.")

    from matplotlib.figure import Figure

    fig = Figure(figsize=(2, 2))
    ax = fig.add_subplot(111)
    ax.plot([0, 1], [0, 1])
    export_figure_png(fig, out_dir / "plot.png")

    excel_path = out_dir / "BeamCoverage_report.xlsx"
    inactive_params = BeamParams(
        nx=2,
        ny=2,
        array_layout="rectangular",
        array_layout_file=str(out_dir / "remembered_array_layout.csv"),
        use_element_pattern=False,
        element_pattern_file=str(out_dir / "remembered_element_pattern.csv"),
        element_near_field_file=str(out_dir / "remembered_element_near_field.csv"),
    )
    inactive_params, inactive_derived = derive_params(inactive_params)
    export_excel_report(
        excel_path,
        params=inactive_params,
        derived=inactive_derived,
        settings=settings,
        timings={"total_s": 0.0},
        figures={"验算图": fig},
        cuts=None,
        current_3d=None,
        uv_pattern=None,
        scan_union=None,
    )
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    text_values: list[str] = []
    for sheet_name in ("报告", "参数"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    text_values.append(value)
    wb.close()
    joined = "\n".join(text_values)
    if (
        "remembered_element_pattern.csv" not in joined
        or "remembered_element_near_field.csv" not in joined
        or "remembered_array_layout.csv" not in joined
        or "未启用" not in joined
        or "各向同性单元" not in joined
        or "参数生成排布" not in joined
        or "近场采样 + 远场外推" not in joined
        or "远场系数扫描并集" not in joined
    ):
        raise RuntimeError("Excel report did not mark inactive remembered imported CSV paths clearly.")

    for path in out_dir.iterdir():
        if path.stat().st_size <= 0:
            raise RuntimeError(f"empty export: {path}")
        print(f"PASS export {path.name} {path.stat().st_size} bytes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
