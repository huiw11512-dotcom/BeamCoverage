from __future__ import annotations

import argparse
import json
import math
import os
import sys
import tempfile
from pathlib import Path

import numpy as np

from app_info import APP_NAME, APP_VERSION, CLI_DESCRIPTION, resource_path
from export.csv_templates import (
    export_current_array_layout_template,
    export_element_pattern_abs_phase_template,
    export_element_pattern_vector_template,
)

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _icon_sizes(icon: object) -> list[list[int]]:
    return [[size.width(), size.height()] for size in icon.availableSizes()]


def _surface_upsample_preserves_nan_gaps(upsample_func: object) -> bool:
    u = np.linspace(-1.0, 1.0, 5)
    v = np.linspace(-1.0, 1.0, 3)
    row = np.array([1.0, 2.0, np.nan, 4.0, 5.0], dtype=float)
    r = np.vstack([row, row + 1.0, row + 2.0])
    u_new, _v_new, r_new = upsample_func(u, v, r, max_u=9, max_v=5)
    gap_columns = [
        int(np.argmin(np.abs(u_new - target)))
        for target in (-0.25, 0.0, 0.25)
    ]
    finite_column = int(np.argmin(np.abs(u_new - -0.75)))
    return bool(
        not np.any(np.isfinite(r_new[:, gap_columns]))
        and np.any(np.isfinite(r_new[:, finite_column]))
    )


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=CLI_DESCRIPTION)
    parser.add_argument("--smoke-test", action="store_true", help="Create the GUI offscreen, run a light calculation, then exit.")
    parser.add_argument("--smoke-output", default="", help="Optional JSON output path for smoke-test status.")
    return parser.parse_args(argv)


def _write_smoke_output(path: str, data: dict[str, object]) -> None:
    if not path:
        return
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(_json_safe(data), indent=2, allow_nan=False), encoding="utf-8")


def _json_safe(value: object) -> object:
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    return value


def _write_smoke_element_pattern(directory: Path) -> Path:
    path = directory / "smoke_element_pattern.csv"
    rows = ["Theta [deg],Phi [deg],dB(GainTotal),Phase(GainTotal)"]
    for theta in range(0, 85, 10):
        for phi in range(-180, 181, 30):
            gain = max(0.05, math.cos(math.radians(theta)) ** 2 * (1.0 - 0.12 * math.sin(math.radians(phi)) ** 2))
            phase = 15.0 * math.sin(math.radians(theta)) * math.cos(math.radians(phi))
            rows.append(f"{theta:.8f},{phi:.8f},{10.0 * math.log10(gain):.8f},{phase:.8f}")
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_radian_element_pattern(directory: Path) -> Path:
    path = directory / "smoke_radian_element_pattern.csv"
    rows = ["Theta [rad],Phi [rad],dB(GainTotal)"]
    for theta_deg in range(0, 85, 10):
        theta = math.radians(theta_deg)
        for phi_deg in range(-180, 181, 30):
            phi = math.radians(phi_deg)
            gain = max(0.05, math.cos(theta) ** 2 * (1.0 - 0.10 * math.cos(phi) ** 2))
            rows.append(f"{theta:.10f},{phi:.10f},{10.0 * math.log10(gain):.8f}")
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_vector_element_pattern(directory: Path) -> Path:
    path = directory / "smoke_vector_element_pattern.csv"
    rows = ["Theta [deg],Phi [deg],Real(ETheta),Imag(ETheta),Real(EPhi),Imag(EPhi)"]
    for theta in range(0, 85, 10):
        for phi in range(-180, 181, 30):
            th = math.radians(theta)
            ph = math.radians(phi)
            amp_theta = max(0.05, math.cos(th) ** 1.15)
            amp_phi = 0.38 * max(0.05, math.cos(th) ** 1.05) * (1.0 + 0.08 * math.cos(ph))
            phase_theta = math.radians(18.0 * math.sin(ph))
            phase_phi = math.radians(-28.0 + 9.0 * math.cos(th))
            etheta = amp_theta * complex(math.cos(phase_theta), math.sin(phase_theta))
            ephi = amp_phi * complex(math.cos(phase_phi), math.sin(phase_phi))
            rows.append(f"{theta:.8f},{phi:.8f},{etheta.real:.8f},{etheta.imag:.8f},{ephi.real:.8f},{ephi.imag:.8f}")
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_metadata_vector_element_pattern(directory: Path) -> Path:
    path = directory / "smoke_metadata_vector_element_pattern.csv"
    rows = [
        "# Solver far-field export",
        "Project,BeamCoverage",
        "Frequency,10 GHz",
        "Setup,LastAdaptive",
        "Theta [deg],Phi [deg],Abs(Theta),Phase(Theta),Abs(Phi),Phase(Phi)",
    ]
    for theta in range(0, 85, 10):
        for phi in range(-180, 181, 30):
            th = math.radians(theta)
            ph = math.radians(phi)
            etheta_abs = max(0.05, math.cos(th) ** 1.12)
            ephi_abs = 0.35 * max(0.05, math.cos(th) ** 1.04) * (1.0 + 0.06 * math.cos(ph))
            etheta_phase = 14.0 * math.sin(ph)
            ephi_phase = -24.0 + 8.0 * math.cos(th)
            rows.append(f"{theta:.8f},{phi:.8f},{etheta_abs:.8f},{etheta_phase:.8f},{ephi_abs:.8f},{ephi_phase:.8f}")
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_multifrequency_element_pattern(directory: Path) -> Path:
    path = directory / "smoke_multifrequency_element_pattern.csv"
    rows = ["Freq,Theta [deg],Phi [deg],Abs(Theta),Phase(Theta),Abs(Phi),Phase(Phi)"]
    for frequency_ghz, scale in ((9.0, 0.72), (11.0, 1.18)):
        frequency_hz = frequency_ghz * 1.0e9
        for theta in range(0, 85, 10):
            for phi in range(-180, 181, 30):
                th = math.radians(theta)
                ph = math.radians(phi)
                etheta_abs = scale * max(0.05, math.cos(th) ** 1.10)
                ephi_abs = 0.34 * scale * max(0.05, math.cos(th) ** 1.02) * (1.0 + 0.05 * math.cos(ph))
                etheta_phase = 11.0 * math.sin(ph)
                ephi_phase = -20.0 + 7.0 * math.cos(th)
                rows.append(
                    f"{frequency_hz:.8f},{theta:.8f},{phi:.8f},{etheta_abs:.8f},{etheta_phase:.8f},"
                    f"{ephi_abs:.8f},{ephi_phase:.8f}"
                )
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_array_layout(directory: Path) -> Path:
    path = directory / "smoke_array_layout.csv"
    rows = ["x_m,y_m,power_w,phase_deg,enabled"]
    coords = [
        (-0.15, -0.10, 1.0e6, 0.0),
        (0.00, -0.12, 0.8e6, 5.0),
        (0.16, -0.08, 1.1e6, -7.0),
        (-0.08, 0.06, 0.9e6, 12.0),
        (0.10, 0.08, 1.2e6, -4.0),
    ]
    for x, y, power, phase in coords:
        rows.append(f"{x:.8f},{y:.8f},{power:.8f},{phase:.8f},1")
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def _write_smoke_overlapping_array_layout(directory: Path) -> Path:
    path = directory / "smoke_overlapping_array_layout.csv"
    rows = [
        "x_m,y_m,power_w",
        "0.00000000,0.00000000,1000.00000000",
        "0.05000000,0.00000000,1000.00000000",
    ]
    path.write_text("\n".join(rows), encoding="utf-8")
    return path


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(sys.argv[1:] if argv is None else argv)
    if args.smoke_test:
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        os.environ.setdefault("BEAMCOVERAGE_DISABLE_SESSION_RESTORE", "1")

    from PySide6.QtGui import QAction, QIcon
    from PySide6.QtWidgets import QApplication

    from gui.main_window import MainWindow
    from gui.plot_panel import _upsample_uv_surface_for_display

    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setApplicationVersion(APP_VERSION)
    icon_path = resource_path("resources", "beamcoverage_icon.png")
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))
    window = MainWindow()
    if args.smoke_test:
        try:
            default_mode_text = window.parameter_panel.calc_mode.currentText()
            default_scan_union_checked = window.parameter_panel.calc_scan_union.isChecked()
            window.parameter_panel.nx.setValue(4)
            window.parameter_panel.ny.setValue(4)
            window.parameter_panel.power.setValue(1.0e4)
            window.parameter_panel.calc_mode.setCurrentIndex(0)
            window.refresh_auto()
            menu_titles = [action.text() for action in window.menuBar().actions()]
            file_action_texts = [action.text() for action in window.findChildren(QAction)]
            menu_localized = bool(
                "文件" in menu_titles
                and "导入阵元坐标 CSV" in file_action_texts
                and "导出当前阵元坐标 CSV" in file_action_texts
                and "导入单元远场方向图文件" in file_action_texts
                and "导出单元远场 Real/Imag 模板" in file_action_texts
                and "导出单元远场幅相模板" in file_action_texts
                and "从单元近场导出远场方向图 CSV" in file_action_texts
            )
            plot_toolbars_ok = bool(
                all(
                    getattr(window.plot_panel.widget(i), "toolbar", None) is not None
                    and getattr(window.plot_panel.widget(i), "canvas", None) is not None
                    and getattr(window.plot_panel.widget(i).toolbar, "canvas", None) is window.plot_panel.widget(i).canvas
                    and len(window.plot_panel.widget(i).toolbar.actions()) >= 5
                    for i in range(window.plot_panel.count())
                    if window.plot_panel.tabText(i) != "更新说明"
                )
            )
            surface_gap_preserve_ok = _surface_upsample_preserves_nan_gaps(_upsample_uv_surface_for_display)
            from core.element_pattern import make_element_pattern
            from core.geometry import BeamParams, base_cache_key, derive_params, sanitized_params

            settings = window.settings
            window.settings = settings.__class__(
                **{
                    **settings.__dict__,
                    "n_alpha_2d": 21,
                    "n_range_2d": 16,
                    "theta_3d_n": 8,
                    "phi_3d_n": 12,
                    "n_range_3d": 12,
                    "n_uv": 31,
                    "scan_union_step_deg": 12.0,
                    "scan_union_theta_n": 8,
                    "scan_union_phi_n": 12,
                    "chunk_size": 256,
                    "scan_block_size": 8,
                }
            )
            window._pending_timer.stop()
            window.calculate_all()
            default_result_array_source_text = window.parameter_panel.results._labels["array_source"].text()
            default_result_active_elements_text = window.parameter_panel.results._labels["active_elements"].text()
            default_result_element_model_text = window.parameter_panel.results._labels["element_model"].text()
            default_current_envelope_method_text = window.parameter_panel.results._labels["current_envelope_method"].text()
            default_scan_union_method_text = window.parameter_panel.results._labels["scan_union_method"].text()
            default_result_metadata_ok = bool(
                default_result_array_source_text == "矩形排布"
                and default_result_active_elements_text == "16"
                and default_result_element_model_text == "矩形解析口径"
                and default_current_envelope_method_text == "近场采样+远场外推"
                and default_scan_union_method_text == "3D远场；2D近场"
            )
            app_icon_sizes = _icon_sizes(app.windowIcon())
            window_icon_sizes = _icon_sizes(window.windowIcon())
            icon_sizes_ok = bool(
                app_icon_sizes
                and window_icon_sizes
                and max(width for width, _height in app_icon_sizes) >= 64
                and max(width for width, _height in window_icon_sizes) >= 64
            )
            default_skip_ok = bool(
                window.cuts
                and window.current_3d
                and window.uv_pattern
                and window.scan_union is None
                and window.last_timings
                and window.last_timings.get("scan_union_skipped")
            )
            current_method_metadata_ok = bool(
                window.current_3d
                and window.current_3d.get("envelopeMethod") == "near_field_sampled_far_field_extrapolated"
                and window.current_3d.get("envelopeMethodLabel") == "近场采样 + 远场外推"
                and window.cuts
                and all(cut.get("envelope_method") == "near_field_sampled_far_field_extrapolated" for cut in window.cuts)
            )
            default_report_figure_keys = sorted(window._report_figures().keys())
            default_report_figures_ok = "扫描并集" not in default_report_figure_keys
            first_timings = window.last_timings or {}

            window.parameter_panel.calc_scan_union.setChecked(True)
            window.calculate_all()
            union_ok = bool(window.scan_union and "Xsurf" in window.scan_union and "maxRange_m" in window.scan_union)
            union_cuts_ok = bool(
                window.scan_union
                and isinstance(window.scan_union.get("unionCuts"), list)
                and len(window.scan_union["unionCuts"]) == 2
                and float(window.scan_union["unionCuts"][0]["phi_cut_deg"]) == 0.0
                and float(window.scan_union["unionCuts"][1]["phi_cut_deg"]) == 90.0
                and np.any(np.isfinite(window.scan_union["unionCuts"][0]["r_env_m"]))
                and np.any(np.isfinite(window.scan_union["unionCuts"][1]["r_env_m"]))
            )
            union_report_figure_keys = sorted(window._report_figures().keys())
            union_report_figures_ok = "扫描并集" in union_report_figure_keys
            union_timings = window.last_timings or {}
            union_shape = list(window.scan_union["Xsurf"].shape) if window.scan_union else None
            union_max_range_m = float(window.scan_union["maxRange_m"]) if window.scan_union else None
            union_cut_sample_counts = (
                [int(np.asarray(cut["r_env_m"]).size) for cut in window.scan_union["unionCuts"]]
                if window.scan_union and isinstance(window.scan_union.get("unionCuts"), list)
                else []
            )
            union_cut_high_resolution_ok = bool(
                union_cut_sample_counts
                and all(count == int(window.settings.n_alpha_2d) for count in union_cut_sample_counts)
            )
            union_method_metadata_ok = bool(
                window.scan_union
                and window.scan_union.get("envelopeMethod") == "far_field_coefficient_union"
                and window.scan_union.get("envelopeMethodLabel") == "远场系数扫描并集"
                and all(cut.get("envelope_method") == "near_field_sampled_far_field_scan_union" for cut in window.scan_union["unionCuts"])
            )
            union_timing_breakdown_ok = bool(
                union_timings
                and not union_timings.get("scan_union_cache_hit")
                and all(
                    np.isfinite(float(union_timings.get(key, float("nan")))) and float(union_timings.get(key, -1.0)) >= 0.0
                    for key in ("scan_union_3d_s", "scan_union_2d_cuts_s", "scan_union_compute_s")
                )
                and window.scan_union
                and isinstance(window.scan_union.get("timings"), dict)
                and all(key in window.scan_union["timings"] for key in ("scan_union_3d_s", "scan_union_2d_cuts_s", "scan_union_compute_s"))
            )
            window.parameter_panel.scan_x.setValue(3.0)
            window.parameter_panel.scan_y.setValue(1.0)
            window._pending_timer.stop()
            window.calculate_all()
            scan_angle_cache_timings = window.last_timings or {}
            scan_angle_cache_reuse_ok = bool(
                scan_angle_cache_timings
                and scan_angle_cache_timings.get("scan_union_cache_hit")
                and scan_angle_cache_timings.get("scan_union_current_overlay")
                and np.isfinite(float(scan_angle_cache_timings.get("scan_union_current_overlay_s", float("nan"))))
                and window.scan_union
                and bool(window.scan_union.get("currentScanOverlayApplied"))
            )
            window.calculate_all()
            scan_angle_full_cache_timings = window.last_timings or {}
            scan_angle_full_cache_reuse_ok = bool(
                scan_angle_full_cache_timings
                and scan_angle_full_cache_timings.get("scan_union_cache_hit")
                and scan_angle_full_cache_timings.get("scan_union_current_cache_hit")
                and scan_angle_full_cache_timings.get("scan_union_current_overlay")
                and window.scan_union
                and bool(window.scan_union.get("currentScanOverlayApplied"))
            )
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_path = Path(tmpdir)
                pattern_path = _write_smoke_element_pattern(tmp_path)
                radian_pattern_path = _write_smoke_radian_element_pattern(tmp_path)
                vector_pattern_path = _write_smoke_vector_element_pattern(tmp_path)
                metadata_vector_pattern_path = _write_smoke_metadata_vector_element_pattern(tmp_path)
                multifrequency_pattern_path = _write_smoke_multifrequency_element_pattern(tmp_path)
                layout_path = _write_smoke_array_layout(tmp_path)
                overlap_layout_path = _write_smoke_overlapping_array_layout(tmp_path)
                exported_array_template_path = tmp_path / "exported_current_array_layout.csv"
                exported_pattern_template_path = tmp_path / "exported_vector_element_pattern_template.csv"
                exported_abs_phase_pattern_template_path = tmp_path / "exported_abs_phase_element_pattern_template.csv"
                project_path_payload = window.project_payload()
                project_path_payload["params"]["array_layout"] = "custom"
                project_path_payload["params"]["array_layout_file"] = str(Path("old_location") / layout_path.name)
                project_path_payload["params"]["element_pattern_file"] = str(Path("old_location") / pattern_path.name)
                window.apply_project_payload(project_path_payload, base_dir=tmp_path)
                window.refresh_auto()
                resolved_project_params = window.parameter_panel.params()
                project_paths_resolved_ok = bool(
                    resolved_project_params.array_layout_file == str(layout_path.resolve())
                    and resolved_project_params.element_pattern_file == str(pattern_path.resolve())
                    and window.derived is not None
                    and window.derived.imported_array_layout
                    and window.elem is not None
                    and window.elem.mode == "table"
                )
                hidden_layout_params = window.parameter_panel.params()
                hidden_layout_params.array_layout = "rectangular"
                hidden_layout_params.array_layout_file = str(layout_path.resolve())
                window.parameter_panel.set_params(hidden_layout_params)
                window.refresh_auto()
                imported_layout_requires_custom_ok = bool(
                    window.params.array_layout == "rectangular"
                    and window.params.array_layout_file == str(layout_path.resolve())
                    and window.derived is not None
                    and not window.derived.imported_array_layout
                )
                inactive_layout_save_relative_ok = bool(
                    window.project_payload_for_file(tmp_path / "inactive_layout.project")["params"].get("array_layout_file")
                    == layout_path.name
                )
                inactive_layout_with_path = BeamParams(
                    nx=2,
                    ny=2,
                    array_layout="rectangular",
                    array_layout_file=str(layout_path.resolve()),
                )
                inactive_layout_without_path = BeamParams(nx=2, ny=2, array_layout="rectangular")
                inactive_layout_params, inactive_layout_derived = derive_params(inactive_layout_with_path)
                inactive_layout_file_ignored_ok = bool(
                    inactive_layout_params.array_layout_file == str(layout_path.resolve())
                    and not inactive_layout_derived.imported_array_layout
                    and base_cache_key(inactive_layout_with_path, window.settings)
                    == base_cache_key(inactive_layout_without_path, window.settings)
                )
                missing_pattern_path = tmp_path / "missing_when_disabled.csv"
                inactive_pattern_with_path = BeamParams(
                    nx=2,
                    ny=2,
                    use_element_pattern=False,
                    element_pattern_file=str(missing_pattern_path),
                )
                inactive_pattern_without_path = BeamParams(nx=2, ny=2, use_element_pattern=False)
                inactive_pattern_params, _ = derive_params(inactive_pattern_with_path)
                inactive_pattern_file_ignored_ok = bool(
                    inactive_pattern_params.element_pattern_file == str(missing_pattern_path)
                    and not inactive_pattern_params.use_element_pattern
                    and base_cache_key(inactive_pattern_with_path, window.settings)
                    == base_cache_key(inactive_pattern_without_path, window.settings)
                )
                settings_cache_key_distinguishes_resolution_ok = bool(
                    base_cache_key(
                        BeamParams(nx=2, ny=2),
                        window.settings.__class__(**{**window.settings.__dict__, "scan_union_theta_n": 9, "scan_union_phi_n": 13}),
                    )
                    != base_cache_key(
                        BeamParams(nx=2, ny=2),
                        window.settings.__class__(**{**window.settings.__dict__, "scan_union_theta_n": 11, "scan_union_phi_n": 15}),
                    )
                )
                string_false_pattern_ignored_ok = bool(
                    not sanitized_params(
                        BeamParams(
                            nx=2,
                            ny=2,
                            use_element_pattern="false",
                            element_pattern_file=str(missing_pattern_path),
                        )
                    ).use_element_pattern
                )
                missing_layout_path = tmp_path / "missing_active_layout.csv"
                missing_import_params = window.parameter_panel.params()
                missing_import_params.array_layout = "custom"
                missing_import_params.array_layout_file = str(missing_layout_path)
                missing_import_params.use_element_pattern = True
                missing_import_params.element_pattern_file = str(missing_pattern_path)
                window.parameter_panel.set_params(missing_import_params)
                window.refresh_auto()
                missing_import_status_labels_ok = bool(
                    window.derived is None
                    and "文件缺失" in window.parameter_panel.array_layout_status.text()
                    and missing_layout_path.name in window.parameter_panel.array_layout_status.text()
                    and "文件缺失" in window.parameter_panel.element_pattern_status.text()
                    and missing_pattern_path.name in window.parameter_panel.element_pattern_status.text()
                )
                missing_project_payload = window.project_payload()
                missing_project_payload["params"]["array_layout"] = "custom"
                missing_project_payload["params"]["array_layout_file"] = str(tmp_path / "missing_project_layout.csv")
                missing_project_payload["params"]["use_element_pattern"] = True
                missing_project_payload["params"]["element_pattern_file"] = str(tmp_path / "missing_project_pattern.csv")
                window.apply_project_payload(missing_project_payload, base_dir=tmp_path)
                window.refresh_auto()
                missing_project_import_status_ok = bool(
                    window.derived is None
                    and "文件缺失" in window.parameter_panel.array_layout_status.text()
                    and "missing_project_layout.csv" in window.parameter_panel.array_layout_status.text()
                    and "文件缺失" in window.parameter_panel.element_pattern_status.text()
                    and "missing_project_pattern.csv" in window.parameter_panel.element_pattern_status.text()
                )
                from export.excel_export import export_excel_report

                inactive_excel_params = BeamParams(
                    nx=2,
                    ny=2,
                    array_layout="rectangular",
                    array_layout_file=str(tmp_path / "remembered_array_layout.csv"),
                    use_element_pattern=False,
                    element_pattern_file=str(tmp_path / "remembered_element_pattern.csv"),
                )
                inactive_excel_params, inactive_excel_derived = derive_params(inactive_excel_params)
                excel_report_path = tmp_path / "inactive_pattern_report.xlsx"
                export_excel_report(
                    excel_report_path,
                    params=inactive_excel_params,
                    derived=inactive_excel_derived,
                    settings=window.settings,
                    timings={"total_s": 0.0},
                    figures={},
                    cuts=None,
                    current_3d=None,
                    uv_pattern=None,
                    scan_union=None,
                )
                from openpyxl import load_workbook

                excel_wb = load_workbook(excel_report_path, read_only=True, data_only=True)
                excel_text_values: list[str] = []
                for sheet_name in ("报告", "参数"):
                    for row in excel_wb[sheet_name].iter_rows(values_only=True):
                        for value in row:
                            if isinstance(value, str):
                                excel_text_values.append(value)
                excel_wb.close()
                excel_text = "\n".join(excel_text_values)
                excel_report_inactive_pattern_label_ok = bool(
                    "remembered_element_pattern.csv" in excel_text
                    and "未启用" in excel_text
                    and "各向同性单元" in excel_text
                )
                excel_report_inactive_layout_label_ok = bool(
                    "remembered_array_layout.csv" in excel_text
                    and "参数生成排布" in excel_text
                )
                window.apply_project_payload(project_path_payload, base_dir=tmp_path)
                window.refresh_auto()
                relative_project_payload = window.project_payload_for_file(tmp_path / "BeamCoverage.project")
                project_save_relative_paths_ok = bool(
                    relative_project_payload["params"].get("array_layout_file") == layout_path.name
                    and relative_project_payload["params"].get("element_pattern_file") == pattern_path.name
                )
                radian_pattern_params = BeamParams(nx=3, ny=3, element_pattern_file=str(radian_pattern_path))
                radian_pattern_params, radian_pattern_derived = derive_params(radian_pattern_params)
                radian_pattern_elem = make_element_pattern(radian_pattern_params, radian_pattern_derived.wavelength_m)
                radian_pattern_loaded_ok = bool(
                    radian_pattern_elem.mode == "table"
                    and 79.5 <= float(radian_pattern_elem.table_theta_max_deg) <= 80.5
                    and not radian_pattern_elem.table_has_phase
                )
                metadata_vector_params = BeamParams(nx=3, ny=3, element_pattern_file=str(metadata_vector_pattern_path))
                metadata_vector_params, metadata_vector_derived = derive_params(metadata_vector_params)
                metadata_vector_elem = make_element_pattern(metadata_vector_params, metadata_vector_derived.wavelength_m)
                metadata_vector_pattern_loaded_ok = bool(
                    metadata_vector_elem.mode == "table"
                    and metadata_vector_elem.table_has_vector_components
                    and metadata_vector_elem.table_has_phase
                    and 79.5 <= float(metadata_vector_elem.table_theta_max_deg) <= 80.5
                )
                multifrequency_low_params = BeamParams(
                    frequency_ghz=9.2,
                    nx=3,
                    ny=3,
                    element_pattern_file=str(multifrequency_pattern_path),
                )
                multifrequency_low_params, multifrequency_low_derived = derive_params(multifrequency_low_params)
                multifrequency_low_elem = make_element_pattern(multifrequency_low_params, multifrequency_low_derived.wavelength_m)
                multifrequency_high_params = BeamParams(
                    frequency_ghz=10.8,
                    nx=3,
                    ny=3,
                    element_pattern_file=str(multifrequency_pattern_path),
                )
                multifrequency_high_params, multifrequency_high_derived = derive_params(multifrequency_high_params)
                multifrequency_high_elem = make_element_pattern(multifrequency_high_params, multifrequency_high_derived.wavelength_m)
                multifrequency_pattern_loaded_ok = bool(
                    multifrequency_low_elem.mode == "table"
                    and multifrequency_high_elem.mode == "table"
                    and multifrequency_low_elem.table_has_vector_components
                    and multifrequency_high_elem.table_has_vector_components
                    and math.isclose(float(multifrequency_low_elem.table_selected_frequency_ghz or -1.0), 9.0, abs_tol=1.0e-12)
                    and math.isclose(float(multifrequency_high_elem.table_selected_frequency_ghz or -1.0), 11.0, abs_tol=1.0e-12)
                    and multifrequency_low_elem.table_point_count == 97
                    and multifrequency_high_elem.table_point_count == 97
                    and not math.isclose(float(multifrequency_low_elem.gain_norm), float(multifrequency_high_elem.gain_norm), rel_tol=1.0e-6)
                )
                export_current_array_layout_template(window.derived, exported_array_template_path)
                export_element_pattern_vector_template(exported_pattern_template_path)
                export_element_pattern_abs_phase_template(exported_abs_phase_pattern_template_path)
                from core.geometry import load_imported_array_layout
                from core.element_pattern import load_imported_element_pattern

                exported_array_template = load_imported_array_layout(exported_array_template_path, window.params.element_power_w)
                exported_pattern_template = load_imported_element_pattern(exported_pattern_template_path)
                exported_abs_phase_pattern_template = load_imported_element_pattern(exported_abs_phase_pattern_template_path)
                export_templates_ok = bool(
                    exported_array_template.point_count == int(window.derived.element_x_m.size)
                    and exported_pattern_template.has_vector_components
                    and exported_pattern_template.point_count > 0
                    and exported_abs_phase_pattern_template.has_vector_components
                    and exported_abs_phase_pattern_template.has_phase
                    and exported_abs_phase_pattern_template.point_count > 0
                )
                imported_params = window.parameter_panel.params()
                imported_params.array_layout = "custom"
                imported_params.array_layout_file = str(layout_path)
                imported_params.element_pattern_file = str(pattern_path)
                window.parameter_panel.set_params(imported_params)
                window.parameter_panel.calc_scan_union.setChecked(False)
                window.refresh_auto()
                imported_pattern_loaded = bool(window.elem and window.elem.mode == "table")
                imported_pattern_has_phase = bool(
                    window.elem
                    and window.elem.table_field is not None
                    and np.nanmax(np.abs(np.imag(window.elem.table_field))) > 1.0e-9
                )
                imported_pattern_status_text = window.parameter_panel.element_pattern_status.text()
                imported_pattern_theta_max_deg = float(window.elem.table_theta_max_deg) if window.elem else None
                imported_pattern_covers_visible_edge = bool(window.elem.table_covers_visible_edge) if window.elem else None
                imported_pattern_status_has_metadata = bool(
                    "含相位/复数场" in imported_pattern_status_text
                    and "点" in imported_pattern_status_text
                    and "θmax=" in imported_pattern_status_text
                    and "未覆盖90°边缘" in imported_pattern_status_text
                    and "Gnorm=" in imported_pattern_status_text
                )
                imported_layout_loaded = bool(window.derived and window.derived.imported_array_layout and window.derived.element_x_m.size == 5)
                imported_array_layout_status_text = window.parameter_panel.array_layout_status.text()
                imported_array_layout_status_has_metadata = bool(
                    "有效单元 5" in imported_array_layout_status_text
                    and "总功率" in imported_array_layout_status_text
                    and "口径" in imported_array_layout_status_text
                )
                imported_result_array_source_text = window.parameter_panel.results._labels["array_source"].text()
                imported_result_active_elements_text = window.parameter_panel.results._labels["active_elements"].text()
                imported_result_element_model_text = window.parameter_panel.results._labels["element_model"].text()
                imported_result_metadata_ok = bool(
                    imported_result_array_source_text == "导入坐标CSV"
                    and imported_result_active_elements_text == "5"
                    and imported_result_element_model_text == "导入远场方向图"
                )
                window.calculate_all()
                imported_pattern_calculated = bool(imported_pattern_loaded and window.cuts and window.uv_pattern and window.current_3d)
                imported_layout_calculated = bool(imported_layout_loaded and imported_pattern_calculated)
                imported_scan_loss_db = float(window.derived.scan_loss_db) if window.derived is not None else None

                vector_params = window.parameter_panel.params()
                vector_params.element_pattern_file = str(vector_pattern_path)
                window.parameter_panel.set_params(vector_params)
                window.parameter_panel.calc_scan_union.setChecked(False)
                window.refresh_auto()
                imported_vector_pattern_loaded = bool(window.elem and window.elem.table_has_vector_components)
                imported_vector_pattern_status_text = window.parameter_panel.element_pattern_status.text()
                imported_vector_pattern_status_has_metadata = bool("矢量Eθ/Eφ" in imported_vector_pattern_status_text)
                imported_vector_result_element_model_text = window.parameter_panel.results._labels["element_model"].text()
                imported_vector_result_metadata_ok = imported_vector_result_element_model_text == "导入远场矢量Eθ/Eφ"
                window.calculate_all()
                imported_vector_pattern_calculated = bool(imported_vector_pattern_loaded and window.cuts and window.uv_pattern and window.current_3d)

                generated_bad_params = window.parameter_panel.params()
                generated_bad_params.array_layout = "rectangular"
                generated_bad_params.array_layout_file = ""
                generated_bad_params.element_shape = "rectangular"
                generated_bad_params.nx = 2
                generated_bad_params.ny = 2
                generated_bad_params.dx_m = 0.10
                generated_bad_params.dy_m = 0.10
                generated_bad_params.ax_m = 0.20
                generated_bad_params.ay_m = 0.10
                window.parameter_panel.set_params(generated_bad_params)
                window.refresh_auto()
                generated_geometry_warning_text = window.parameter_panel.geometry_warning.text()
                generated_geometry_calculate_disabled = not window.parameter_panel.calculate_button.isEnabled()
                generated_geometry_overlap_rejected = bool(
                    window.derived is None
                    and "参数生成阵列" in generated_geometry_warning_text
                    and "口径重叠" in generated_geometry_warning_text
                    and generated_geometry_calculate_disabled
                )
                try:
                    _, single_axis_derived = derive_params(
                        BeamParams(nx=1, ny=2, dx_m=0.10, dy_m=0.10, ax_m=0.20, ay_m=0.10)
                    )
                    generated_geometry_single_axis_ok = bool(single_axis_derived.element_x_m.size == 2)
                except ValueError:
                    generated_geometry_single_axis_ok = False

                overlap_params = window.parameter_panel.params()
                overlap_params.array_layout = "custom"
                overlap_params.array_layout_file = str(overlap_layout_path)
                overlap_params.element_shape = "rectangular"
                overlap_params.ax_m = 0.10
                overlap_params.ay_m = 0.05
                window.parameter_panel.set_params(overlap_params)
                window.refresh_auto()
                imported_array_overlap_warning_text = window.parameter_panel.geometry_warning.text()
                imported_array_overlap_calculate_disabled = not window.parameter_panel.calculate_button.isEnabled()
                imported_array_overlap_rejected = bool(
                    window.derived is None
                    and "口径重叠" in imported_array_overlap_warning_text
                    and "口径重叠" in window.statusBar().currentMessage()
                    and imported_array_overlap_calculate_disabled
                )

            ok = bool(
                default_skip_ok
                and current_method_metadata_ok
                and union_ok
                and union_cuts_ok
                and union_cut_high_resolution_ok
                and union_method_metadata_ok
                and union_timing_breakdown_ok
                and scan_angle_cache_reuse_ok
                and scan_angle_full_cache_reuse_ok
                and settings_cache_key_distinguishes_resolution_ok
                and imported_pattern_calculated
                and imported_layout_calculated
                and imported_pattern_status_has_metadata
                and imported_vector_pattern_calculated
                and imported_vector_pattern_status_has_metadata
                and generated_geometry_overlap_rejected
                and generated_geometry_single_axis_ok
                and imported_array_overlap_rejected
                and imported_array_layout_status_has_metadata
                and default_result_metadata_ok
                and imported_result_metadata_ok
                and imported_vector_result_metadata_ok
                and export_templates_ok
                and menu_localized
                and plot_toolbars_ok
                and surface_gap_preserve_ok
                and default_report_figures_ok
                and union_report_figures_ok
                and project_paths_resolved_ok
                and project_save_relative_paths_ok
                and radian_pattern_loaded_ok
                and metadata_vector_pattern_loaded_ok
                and multifrequency_pattern_loaded_ok
                and imported_layout_requires_custom_ok
                and inactive_layout_save_relative_ok
                and inactive_layout_file_ignored_ok
                and inactive_pattern_file_ignored_ok
                and string_false_pattern_ignored_ok
                and missing_import_status_labels_ok
                and missing_project_import_status_ok
                and excel_report_inactive_pattern_label_ok
                and excel_report_inactive_layout_label_ok
                and icon_sizes_ok
            )
            payload = {
                "ok": ok,
                "title": window.windowTitle(),
                "app_name": APP_NAME,
                "app_version": APP_VERSION,
                "menu_titles": menu_titles,
                "file_action_texts": file_action_texts,
                "menu_localized": menu_localized,
                "plot_toolbars_ok": plot_toolbars_ok,
                "surface_gap_preserve_ok": surface_gap_preserve_ok,
                "app_icon_loaded": not app.windowIcon().isNull(),
                "window_icon_loaded": not window.windowIcon().isNull(),
                "app_icon_sizes": app_icon_sizes,
                "window_icon_sizes": window_icon_sizes,
                "icon_sizes_ok": icon_sizes_ok,
                "tabs": [window.plot_panel.tabText(i) for i in range(window.plot_panel.count())],
                "default_calc_mode": default_mode_text,
                "default_scan_union_checked": default_scan_union_checked,
                "default_scan_union_skipped": default_skip_ok,
                "current_method_metadata_ok": current_method_metadata_ok,
                "default_report_figure_keys": default_report_figure_keys,
                "default_report_figures_ok": default_report_figures_ok,
                "union_report_figure_keys": union_report_figure_keys,
                "union_report_figures_ok": union_report_figures_ok,
                "union_cuts_ok": union_cuts_ok,
                "union_cut_sample_counts": union_cut_sample_counts,
                "union_cut_high_resolution_ok": union_cut_high_resolution_ok,
                "union_method_metadata_ok": union_method_metadata_ok,
                "union_timing_breakdown_ok": union_timing_breakdown_ok,
                "scan_angle_cache_reuse_ok": scan_angle_cache_reuse_ok,
                "scan_angle_full_cache_reuse_ok": scan_angle_full_cache_reuse_ok,
                "timings_after_scan_angle_change": scan_angle_cache_timings,
                "timings_after_repeated_scan_angle": scan_angle_full_cache_timings,
                "settings_cache_key_distinguishes_resolution_ok": settings_cache_key_distinguishes_resolution_ok,
                "project_paths_resolved_ok": project_paths_resolved_ok,
                "project_save_relative_paths_ok": project_save_relative_paths_ok,
                "radian_pattern_loaded_ok": radian_pattern_loaded_ok,
                "radian_pattern_theta_max_deg": float(radian_pattern_elem.table_theta_max_deg),
                "metadata_vector_pattern_loaded_ok": metadata_vector_pattern_loaded_ok,
                "multifrequency_pattern_loaded_ok": multifrequency_pattern_loaded_ok,
                "multifrequency_selected_frequencies_ghz": [
                    multifrequency_low_elem.table_selected_frequency_ghz,
                    multifrequency_high_elem.table_selected_frequency_ghz,
                ],
                "imported_layout_requires_custom_ok": imported_layout_requires_custom_ok,
                "inactive_layout_save_relative_ok": inactive_layout_save_relative_ok,
                "inactive_layout_file_ignored_ok": inactive_layout_file_ignored_ok,
                "inactive_pattern_file_ignored_ok": inactive_pattern_file_ignored_ok,
                "string_false_pattern_ignored_ok": string_false_pattern_ignored_ok,
                "missing_import_status_labels_ok": missing_import_status_labels_ok,
                "missing_project_import_status_ok": missing_project_import_status_ok,
                "excel_report_inactive_pattern_label_ok": excel_report_inactive_pattern_label_ok,
                "excel_report_inactive_layout_label_ok": excel_report_inactive_layout_label_ok,
                "generated_geometry_overlap_rejected": generated_geometry_overlap_rejected,
                "generated_geometry_calculate_disabled": generated_geometry_calculate_disabled,
                "generated_geometry_single_axis_ok": generated_geometry_single_axis_ok,
                "generated_geometry_warning_text": generated_geometry_warning_text,
                "default_result_array_source_text": default_result_array_source_text,
                "default_result_active_elements_text": default_result_active_elements_text,
                "default_result_element_model_text": default_result_element_model_text,
                "default_current_envelope_method_text": default_current_envelope_method_text,
                "default_scan_union_method_text": default_scan_union_method_text,
                "default_result_metadata_ok": default_result_metadata_ok,
                "export_templates_ok": export_templates_ok,
                "exported_array_template_point_count": int(exported_array_template.point_count),
                "exported_element_pattern_template_vector": bool(exported_pattern_template.has_vector_components),
                "exported_element_pattern_abs_phase_template_vector": bool(exported_abs_phase_pattern_template.has_vector_components),
                "exported_element_pattern_abs_phase_template_has_phase": bool(exported_abs_phase_pattern_template.has_phase),
                "timings_default": first_timings,
                "timings_with_union": union_timings,
                "scan_union_shape": union_shape,
                "scan_union_max_range_m": union_max_range_m,
                "imported_element_pattern_loaded": imported_pattern_loaded,
                "imported_element_pattern_calculated": imported_pattern_calculated,
                "imported_element_pattern_has_phase": imported_pattern_has_phase,
                "imported_element_pattern_theta_max_deg": imported_pattern_theta_max_deg,
                "imported_element_pattern_covers_visible_edge": imported_pattern_covers_visible_edge,
                "imported_element_pattern_status_text": imported_pattern_status_text,
                "imported_element_pattern_status_has_metadata": imported_pattern_status_has_metadata,
                "imported_vector_element_pattern_loaded": imported_vector_pattern_loaded,
                "imported_vector_element_pattern_calculated": imported_vector_pattern_calculated,
                "imported_vector_element_pattern_status_text": imported_vector_pattern_status_text,
                "imported_vector_element_pattern_status_has_metadata": imported_vector_pattern_status_has_metadata,
                "imported_array_layout_loaded": imported_layout_loaded,
                "imported_array_layout_calculated": imported_layout_calculated,
                "imported_array_layout_status_text": imported_array_layout_status_text,
                "imported_array_layout_status_has_metadata": imported_array_layout_status_has_metadata,
                "imported_result_array_source_text": imported_result_array_source_text,
                "imported_result_active_elements_text": imported_result_active_elements_text,
                "imported_result_element_model_text": imported_result_element_model_text,
                "imported_result_metadata_ok": imported_result_metadata_ok,
                "imported_array_overlap_rejected": imported_array_overlap_rejected,
                "imported_array_overlap_calculate_disabled": imported_array_overlap_calculate_disabled,
                "imported_array_overlap_warning_text": imported_array_overlap_warning_text,
                "imported_vector_result_element_model_text": imported_vector_result_element_model_text,
                "imported_vector_result_metadata_ok": imported_vector_result_metadata_ok,
                "imported_scan_loss_db": imported_scan_loss_db,
            }
            _write_smoke_output(args.smoke_output, payload)
            window.close()
            app.quit()
            return 0 if ok else 2
        except Exception as exc:
            _write_smoke_output(args.smoke_output, {"ok": False, "error": str(exc)})
            return 1

    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
