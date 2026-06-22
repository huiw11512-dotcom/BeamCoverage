from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.aperture_shapes import shape_label
from core.geometry import BeamParams, DerivedParams, ModeSettings


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(bold=True)


def export_excel_report(
    path: str | Path,
    *,
    params: BeamParams,
    derived: DerivedParams,
    settings: ModeSettings,
    timings: dict[str, Any] | None,
    figures: dict[str, Figure],
    cuts: list[dict[str, Any]] | None,
    current_3d: dict[str, Any] | None,
    uv_pattern: dict[str, Any] | None,
    scan_union: dict[str, Any] | None,
) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "报告"
    _write_report_sheet(ws, params, derived, settings, timings)
    _write_key_value_sheet(wb.create_sheet("参数"), _parameter_rows(params))
    _write_key_value_sheet(wb.create_sheet("自动计算结果"), _derived_rows(derived, timings, settings))
    _write_plot_sheet(wb.create_sheet("图形"), figures)
    _write_2d_cuts_sheet(wb.create_sheet("二维切面数据"), cuts, params.s0_w_cm2)
    _write_columns_sheet(wb.create_sheet("三维包络数据"), _current_3d_columns(current_3d, params.s0_w_cm2), "当前三维包络未计算。")
    _write_columns_sheet(wb.create_sheet("方向图数据"), _uv_columns(uv_pattern), "方向图未计算。")
    _write_columns_sheet(wb.create_sheet("扫描并集数据"), _scan_union_columns(scan_union), "扫描并集未计算。")

    for sheet in wb.worksheets:
        _freeze_and_style(sheet)
    wb.save(target)


def _write_report_sheet(ws, params: BeamParams, derived: DerivedParams, settings: ModeSettings, timings: dict[str, Any] | None) -> None:
    ws["A1"] = "BeamCoverage 计算报告"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("频率", params.frequency_ghz, "GHz"),
        ("阵列规模", f"{params.nx} x {params.ny}"),
        ("阵列排布", shape_label(getattr(params, "array_layout", "rectangular"), kind="generic", default=str(params.array_layout))),
        ("阵元坐标文件", _array_layout_file_label(params)),
        ("单元口径", shape_label(getattr(params, "element_shape", "rectangular"), kind="generic", default=str(params.element_shape))),
        ("单元远场方向图文件", _element_pattern_file_label(params)),
        ("单元近场文件", _element_near_field_file_label(params)),
        ("单元间距", f"dx={params.dx_m:g} m, dy={params.dy_m:g} m"),
        ("单元尺寸", f"ax={params.ax_m:g} m, ay={params.ay_m:g} m"),
        ("扫描角", f"x-z={params.scan_x_deg:g} deg, y-z={params.scan_y_deg:g} deg"),
        ("当前 theta / phi", f"{derived.theta_deg:.6g} deg / {derived.phi_deg:.6g} deg"),
        ("计算模式", settings.name),
        ("总输入功率", derived.total_input_power_w, "W"),
        ("Rff", derived.rff_m, "m"),
        ("HPBWx / HPBWy", f"{derived.hpbw_x_deg:.6g} deg / {derived.hpbw_y_deg:.6g} deg"),
        ("二维/当前三维包络方法", "近场采样 + 远场外推"),
        ("扫描并集包络方法", "远场系数扫描并集"),
        ("总耗时", None if not timings else timings.get("total_s"), "s"),
        ("扫描并集3D预览耗时", None if not timings else timings.get("scan_union_3d_s"), "s"),
        ("扫描并集2D切面耗时", None if not timings else timings.get("scan_union_2d_cuts_s"), "s"),
        ("扫描并集当前角增量耗时", None if not timings else timings.get("scan_union_current_overlay_s"), "s"),
        ("扫描并集当前角缓存命中", None if not timings else timings.get("scan_union_current_cache_hit"), ""),
    ]
    _write_rows(ws, 3, ["项目", "值", "单位"], rows)


def _write_key_value_sheet(ws, rows: list[tuple[Any, ...]]) -> None:
    _write_rows(ws, 1, ["项目", "值", "单位"], rows)


def _write_rows(ws, start_row: int, headers: list[str], rows: list[tuple[Any, ...]]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, _excel_value(value))


def _write_plot_sheet(ws, figures: dict[str, Figure]) -> None:
    ws["A1"] = "图形"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    row = 3
    image_buffers: list[BytesIO] = []
    for title, figure in figures.items():
        ws.cell(row, 1, title).font = HEADER_FONT
        buffer = BytesIO()
        figure.savefig(buffer, format="png", dpi=130, bbox_inches="tight")
        buffer.seek(0)
        image = XLImage(buffer)
        max_width = 900
        if image.width > max_width:
            ratio = max_width / image.width
            image.width = int(image.width * ratio)
            image.height = int(image.height * ratio)
        ws.add_image(image, f"A{row + 1}")
        image_buffers.append(buffer)
        row += max(18, int(image.height / 18) + 4)
    ws._beamcoverage_image_buffers = image_buffers


def _write_2d_cuts_sheet(ws, cuts: list[dict[str, Any]] | None, s0_w_cm2: float) -> None:
    if not cuts:
        ws["A1"] = "二维切面未计算。"
        return
    rows: list[dict[str, Any]] = []
    for idx, cut in enumerate(cuts, start=1):
        alpha = _as_1d(cut["alpha_deg"])
        n = alpha.size
        columns = {
            "cut_id": np.full(n, idx),
            "cut_name": np.full(n, cut["name"], dtype=object),
            "phi_cut_deg": np.full(n, cut["phi_cut_deg"]),
            "alpha_deg": alpha,
            "r_env_m": _as_1d(cut["r_env_m"]),
            "rho_m": _as_1d(cut["rho_env_m"]),
            "x_m": _as_1d(cut["x_env_m"]),
            "y_m": _as_1d(cut["y_env_m"]),
            "z_m": _as_1d(cut["z_env_m"]),
            "has_envelope": _as_1d(cut["has_envelope"]),
            "far_field_extrapolated": _as_1d(cut["far_extended"]),
            "range_clipped": _as_1d(cut["clipped"]),
            "S0_Wcm2": np.full(n, s0_w_cm2),
        }
        for row_idx in range(n):
            rows.append({key: value[row_idx] for key, value in columns.items()})
    _write_dict_rows(ws, rows)


def _write_columns_sheet(ws, columns: dict[str, Any] | None, empty_message: str) -> None:
    if not columns:
        ws["A1"] = empty_message
        return
    expanded = _expand_columns(columns)
    headers = list(expanded.keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    n_rows = max((arr.size for arr in expanded.values()), default=0)
    for row_idx in range(n_rows):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx + 2, col_idx, _excel_value(expanded[header][row_idx]))


def _write_dict_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "无数据。"
        return
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, _excel_value(row[header]))


def _parameter_rows(params: BeamParams) -> list[tuple[Any, ...]]:
    labels = {
        "frequency_ghz": ("频率 f", "GHz"),
        "nx": ("Nx", ""),
        "ny": ("Ny", ""),
        "dx_m": ("dx", "m"),
        "dy_m": ("dy", "m"),
        "ax_m": ("ax", "m"),
        "ay_m": ("ay", "m"),
        "array_layout": ("阵列排布", ""),
        "array_layout_file": ("阵元坐标文件", ""),
        "element_shape": ("单元口径", ""),
        "efficiency": ("效率 eta", ""),
        "element_power_w": ("每单元功率", "W"),
        "s0_w_cm2": ("S0", "W/cm2"),
        "scan_x_deg": ("x-z切面扫描角", "deg"),
        "scan_y_deg": ("y-z切面扫描角", "deg"),
        "scan_limit_mode": ("最大扫描角模式", ""),
        "manual_scan_limit_x_deg": ("手动X最大扫描角", "deg"),
        "manual_scan_limit_y_deg": ("手动Y最大扫描角", "deg"),
        "calc_mode": ("计算精度", ""),
        "use_element_pattern": ("使用单元远场方向图", ""),
        "element_pattern_file": ("单元远场方向图文件", ""),
        "element_near_field_file": ("单元近场文件", ""),
    }
    data = params.to_dict()
    rows: list[tuple[Any, ...]] = []
    for key in labels:
        value = data[key]
        if key in ("array_layout", "element_shape"):
            value = shape_label(str(value), kind="generic", default=str(value))
        if key == "array_layout_file":
            value = _array_layout_file_label(params)
        if key == "element_pattern_file":
            value = _element_pattern_file_label(params)
        if key == "element_near_field_file":
            value = _element_near_field_file_label(params)
        rows.append((labels[key][0], value, labels[key][1]))
    return rows


def _array_layout_file_label(params: BeamParams) -> str:
    path = str(getattr(params, "array_layout_file", "") or "")
    if not path:
        return "参数生成"
    name = Path(path).name
    if str(getattr(params, "array_layout", "rectangular")) == "custom":
        return f"CSV: {name}（已启用）"
    return f"CSV: {name}（未启用，当前为参数生成排布）"


def _element_pattern_file_label(params: BeamParams) -> str:
    path = str(getattr(params, "element_pattern_file", "") or "")
    enabled = _bool_param(getattr(params, "use_element_pattern", True))
    if not enabled:
        if path:
            return f"CSV: {Path(path).name}（未启用，当前为各向同性单元）"
        return "各向同性单元（未使用单元远场方向图）"
    if path:
        return f"CSV: {Path(path).name}（已启用）"
    return "内置解析模型"


def _element_near_field_file_label(params: BeamParams) -> str:
    path = str(getattr(params, "element_near_field_file", "") or "")
    if path:
        return f"CSV: {Path(path).name}（实验接口，当前不参与主包络计算）"
    return "未导入"


def _bool_param(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"0", "false", "no", "off", "disabled", "禁用", "否"}:
        return False
    if text in {"1", "true", "yes", "on", "enabled", "启用", "是"}:
        return True
    return bool(value)


def _derived_rows(derived: DerivedParams, timings: dict[str, Any] | None, settings: ModeSettings) -> list[tuple[Any, ...]]:
    rows = [
        ("lambda", derived.wavelength_m, "m"),
        ("Dx", derived.dx_aperture_m, "m"),
        ("Dy", derived.dy_aperture_m, "m"),
        ("Rff", derived.rff_m, "m"),
        ("HPBWx", derived.hpbw_x_deg, "deg"),
        ("HPBWy", derived.hpbw_y_deg, "deg"),
        ("经验最大扫描角X", derived.scan_limit_x_deg, "deg"),
        ("经验最大扫描角Y", derived.scan_limit_y_deg, "deg"),
        ("当前theta", derived.theta_deg, "deg"),
        ("当前phi", derived.phi_deg, "deg"),
        ("当前扫描损失", derived.scan_loss_db, "dB"),
        ("相邻单元相位差X", derived.phase_step_x_deg, "deg"),
        ("相邻单元相位差Y", derived.phase_step_y_deg, "deg"),
        ("总输入功率", derived.total_input_power_w, "W"),
        ("计算模式", settings.name, ""),
        ("二维/当前三维包络方法", "近场采样 + 远场外推", ""),
        ("扫描并集包络方法", "远场系数扫描并集", ""),
    ]
    if timings:
        rows.extend(
            [
                ("二维切面耗时", timings.get("cuts_s"), "s"),
                ("三维包络耗时", timings.get("current_3d_s"), "s"),
                ("方向图耗时", timings.get("uv_s"), "s"),
                ("扫描并集耗时", timings.get("scan_union_s"), "s"),
                ("扫描并集3D预览耗时", timings.get("scan_union_3d_s"), "s"),
                ("扫描并集2D切面耗时", timings.get("scan_union_2d_cuts_s"), "s"),
                ("扫描并集缓存命中", timings.get("scan_union_cache_hit"), ""),
                ("扫描并集当前角缓存命中", timings.get("scan_union_current_cache_hit"), ""),
                ("扫描并集当前角增量", timings.get("scan_union_current_overlay"), ""),
                ("扫描并集当前角增量耗时", timings.get("scan_union_current_overlay_s"), "s"),
                ("总耗时", timings.get("total_s"), "s"),
            ]
        )
    return rows


def _current_3d_columns(envelope: dict[str, Any] | None, s0_w_cm2: float) -> dict[str, Any] | None:
    if not envelope:
        return None
    return {
        "theta_deg": envelope["THETA"],
        "phi_deg": envelope["PHI"],
        "x_m": envelope["Xsurf"],
        "y_m": envelope["Ysurf"],
        "z_m": envelope["Zsurf"],
        "r_env_m": envelope["r_env_m"],
        "has_envelope": envelope["has_envelope"],
        "far_field_extrapolated": envelope["far_extended"],
        "range_clipped": envelope["clipped"],
        "S0_Wcm2": np.full(np.asarray(envelope["THETA"]).shape, s0_w_cm2),
        "envelope_method": envelope.get("envelopeMethod", ""),
        "envelope_method_label": envelope.get("envelopeMethodLabel", ""),
    }


def _uv_columns(uv: dict[str, Any] | None) -> dict[str, Any] | None:
    if not uv:
        return None
    theta_x = np.degrees(np.arcsin(np.clip(np.asarray(uv["U"], dtype=float), -1.0, 1.0)))
    theta_y = np.degrees(np.arcsin(np.clip(np.asarray(uv["V"], dtype=float), -1.0, 1.0)))
    return {
        "u": uv["U"],
        "v": uv["V"],
        "w": uv["W"],
        "theta_x_deg": theta_x,
        "theta_y_deg": theta_y,
        "visible": uv["visible"],
        "pattern_db": uv["pattern_db"],
    }


def _scan_union_columns(info: dict[str, Any] | None) -> dict[str, Any] | None:
    if not info:
        return None
    return {
        "theta_deg": info["THETA"],
        "phi_deg": info["PHI"],
        "x_m": info["Xsurf"],
        "y_m": info["Ysurf"],
        "z_m": info["Zsurf"],
        "r_env_m": info["Rsurf"],
        "best_scan_x_deg": info["bestScanX_deg"],
        "best_scan_y_deg": info["bestScanY_deg"],
        "max_coef": info["maxCoef"],
        "envelope_method": info.get("envelopeMethod", ""),
        "envelope_method_label": info.get("envelopeMethodLabel", ""),
    }


def _expand_columns(columns: dict[str, Any]) -> dict[str, np.ndarray]:
    arrays = {key: _as_1d(value) for key, value in columns.items()}
    n_rows = max((arr.size for arr in arrays.values()), default=0)
    expanded: dict[str, np.ndarray] = {}
    for key, arr in arrays.items():
        if arr.size == n_rows:
            expanded[key] = arr
        elif arr.size == 1:
            expanded[key] = np.full(n_rows, arr.item(), dtype=object)
        else:
            raise ValueError(f"Excel column {key!r} has length {arr.size}; expected {n_rows}.")
    return expanded


def _as_1d(value: Any) -> np.ndarray:
    return np.asarray(value).ravel()


def _excel_value(value: Any) -> Any:
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
        return None
    if isinstance(value, (list, tuple, dict)):
        return str(value)
    return value


def _freeze_and_style(ws) -> None:
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    for col_idx in range(1, min(ws.max_column, 12) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(12, _column_width(ws, col_idx)), 34)


def _column_width(ws, col_idx: int) -> int:
    width = 12
    for row_idx in range(1, min(ws.max_row, 80) + 1):
        value = ws.cell(row_idx, col_idx).value
        if value is not None:
            width = max(width, len(str(value)) + 2)
    return width
